"""Link Excel purchase references to Odoo purchase orders.

Reads purchase references from Excel workbooks, resolves matching records in
Odoo purchase orders with workbook-specific lookup rules, and adds hyperlinks
back to the Excel cells.

By default runs in dry-run mode; pass --apply to update the workbooks.
"""

from __future__ import annotations

import argparse
import copy
import ctypes
import csv
import datetime as dt
import getpass
import hashlib
import os
import posixpath
import re
import shutil
import sys
import tempfile
import threading
import time
import zipfile
import xml.etree.ElementTree as ET
import xmlrpc.client
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import contextmanager
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Any
from urllib.parse import urlencode, urlparse
from xml.sax.saxutils import quoteattr

import pythoncom

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import win32com.client  # type: ignore[import-not-found]
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "pywin32 is required for Excel COM automation. Install it with: "
        "python -m pip install pywin32"
    ) from exc

try:
    import xlwings as xw  # type: ignore[import-not-found]
except ImportError:
    xw = None

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DEFAULT_ODOO_URL = "https://sphe.cloudoo.ma"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
DEFAULT_EXCEL_SESSION_BACKEND = "pywin32"
DEFAULT_EXCEL_SAVE_DEBOUNCE_SECONDS = 1
PERFORMANCE_MODE_SILENT = "silent"
PERFORMANCE_MODE_LIVE = "live"
DEFAULT_ODOO_MAX_WORKERS = 3
LOCAL_WORKBOOK_FILE_NAME = "excel facture achats local.xlsx"
ETRANGER_WORKBOOK_FILE_NAME = "tracking achats etranger (1).xlsx"
LOOKUP_MODE_PARTNER_REF = "partner_ref"
LOOKUP_MODE_COMMAND_REF = "command_ref"
LOOKUP_MODE_TOTAL_AMOUNT = "total_amount"
LOOKUP_MODE_GLOBAL_TEXT = "global_text"
WORKBOOK_SLOT_ACHATS_LOCAL = "achats_local"
WORKBOOK_SLOT_ACHATS_ETRANGER = "achats_etranger"
WORKBOOK_SLOT_SELLER_PREVIOUS = "seller_previous"
DEFAULT_AMOUNT_TOLERANCES = (0.01, 0.05, 0.50)
ODOO_RPC_MAX_ATTEMPTS = 3
ODOO_RPC_RETRY_BASE_SECONDS = 0.75
GLOBAL_SEARCH_FIELD_GROUP_SIZE = 7
GLOBAL_SEARCH_REF_CHUNK_SIZE = 80
GLOBAL_SEARCH_RECORD_LIMIT = 500
GLOBAL_SEARCH_CONTAINS_LIMIT = 5
BACKUP_RUNS_DIR_NAME = "run-backups"
BACKUP_ORIGINALS_DIR_NAME = "original-snapshots"
OOXML_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OOXML_MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
OOXML_DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
OOXML_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OOXML_HYPERLINK_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink"
OOXML_KNOWN_IGNORABLE_NAMESPACES = {
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "x16r2": "http://schemas.microsoft.com/office/spreadsheetml/2015/02/main",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
}
OOXML_HYPERLINK_INSERT_BEFORE_TAGS = (
    "printOptions", "pageMargins", "pageSetup", "headerFooter",
    "rowBreaks", "colBreaks", "customProperties", "cellWatches",
    "ignoredErrors", "smartTags", "drawing", "legacyDrawing",
    "legacyDrawingHF", "picture", "oleObjects", "controls",
    "webPublishItems", "tableParts", "extLst",
)

ET.register_namespace("", OOXML_MAIN_NS)
ET.register_namespace("mc", OOXML_MC_NS)
ET.register_namespace("r", OOXML_DOC_REL_NS)
for _ooxml_prefix, _ooxml_uri in OOXML_KNOWN_IGNORABLE_NAMESPACES.items():
    ET.register_namespace(_ooxml_prefix, _ooxml_uri)

# Normalized header sets used by the ACHATS and seller workflows.
LOCAL_HEADER_VARIANTS = {"n\u00b0facture", "n\u00b0 facture", "nfacture", "n facture"}
LOCAL_COMMAND_HEADER_VARIANTS = {
    "n commandes",
    "n commande",
    "n\u00b0 commandes",
    "n\u00b0 commande",
    "ncommandes",
    "ncommande",
}
ETRANGER_HEADER_VARIANTS = LOCAL_COMMAND_HEADER_VARIANTS
ETRANGER_TOTAL_HEADER_VARIANTS = {"mtt de facture"}
ALL_HEADER_VARIANTS = LOCAL_HEADER_VARIANTS | ETRANGER_HEADER_VARIANTS

PURCHASE_ORDER_FIELDS = [
    "name", "partner_ref", "partner_id", "state", "date_order",
    "amount_total", "origin",
]

GENERIC_RECORD_FIELDS = ["display_name"]
GLOBAL_SEARCH_TEXT_TYPES = {"char", "text", "html", "selection", "reference"}
GLOBAL_SEARCH_EXCLUDED_FIELDS = {
    "access_token",
    "activity_type_icon",
    "website_url",
}
GLOBAL_SEARCH_MODEL_CANDIDATES = [
    "purchase.order", "purchase.order.line", "purchase.requisition", "purchase.requisition.line",
    "account.move", "account.move.line", "account.payment", "account.payment.register",
    "account.bank.statement", "account.bank.statement.line", "account.analytic.line",
    "sale.order", "sale.order.line", "sale.subscription", "sale.subscription.line",
    "stock.picking", "stock.move", "stock.move.line", "stock.lot", "stock.quant",
    "stock.valuation.layer", "stock.scrap", "stock.warehouse.orderpoint", "stock.warehouse",
    "stock.location", "product.template", "product.product", "product.supplierinfo",
    "product.category", "product.pricelist", "product.pricelist.item", "res.partner",
    "res.company", "res.users", "res.country", "res.currency", "crm.lead", "project.project",
    "project.task", "helpdesk.ticket", "maintenance.request", "maintenance.equipment",
    "fleet.vehicle", "fleet.vehicle.log.services", "mrp.production", "mrp.workorder",
    "mrp.bom", "mrp.bom.line", "repair.order", "repair.line", "documents.document",
    "document.document", "ir.attachment", "mail.message", "mail.activity", "mail.followers",
    "calendar.event", "hr.expense", "hr.expense.sheet", "hr.employee", "hr.department",
    "account.journal", "account.account", "account.tax", "account.full.reconcile",
    "account.partial.reconcile", "account.payment.term", "account.incoterms",
    "account.edi.document", "uom.uom", "uom.category", "delivery.carrier", "website.sale.order",
    "pos.order", "pos.order.line", "pos.payment", "quality.check", "quality.alert",
    "quality.point", "sign.request", "sign.request.item",
]
GLOBAL_SEARCH_PRIORITY_FIELDS = {
    "account.move": ("ref", "payment_reference", "name", "display_name", "invoice_origin", "x_studio_n_dossier"),
    "purchase.order": ("partner_ref", "name", "display_name", "origin", "x_studio_n_dossier", "x_studio_description_dossier"),
    "account.move.line": ("ref", "move_name", "name", "display_name", "invoice_origin", "matching_number"),
    "stock.picking": ("origin", "name", "display_name", "carrier_tracking_ref"),
    "stock.move": ("origin", "reference", "name", "display_name"),
    "sale.order": ("client_order_ref", "name", "display_name", "origin", "reference"),
    "ir.attachment": ("display_name", "name", "res_model"),
}
GLOBAL_FAST_EXACT_FIELDS = {
    "account.move": ("ref", "payment_reference", "name", "invoice_origin", "x_studio_n_dossier"),
    "purchase.order": ("partner_ref", "name", "origin", "x_studio_n_dossier", "x_studio_description_dossier"),
    "account.move.line": ("ref", "move_name", "name", "invoice_origin", "matching_number"),
    "stock.picking": ("origin", "name", "carrier_tracking_ref"),
    "stock.move": ("origin", "reference", "name"),
    "ir.attachment": ("name", "display_name"),
}
GLOBAL_FAST_CONTAINS_FIELDS = {
    "account.move": ("ref", "payment_reference", "display_name", "invoice_origin"),
    "purchase.order": ("partner_ref", "name", "display_name", "origin"),
    "ir.attachment": ("name", "display_name"),
    "stock.picking": ("origin", "name", "carrier_tracking_ref"),
}
GLOBAL_SEARCH_MODEL_ORDER = tuple(dict.fromkeys(
    list(GLOBAL_SEARCH_PRIORITY_FIELDS)
    + GLOBAL_SEARCH_MODEL_CANDIDATES
))
_GLOBAL_SEARCH_FIELDS_CACHE: dict[tuple[str, str, str], dict[str, list[str]]] = {}
_GLOBAL_FAST_FIELDS_CACHE: dict[tuple[str, str, str, str], dict[str, list[str]]] = {}

REPORT_COLUMNS = (
    "sheet", "cell", "reference", "status", "source_model",
    "matched_field", "record_id", "record_name", "ref_value",
    "state", "vendor", "amount", "url", "note",
)

LIVE_UPDATE_READY_STATES = {"open_writable"}
LIVE_UPDATE_WAIT_STATES = {
    "open_read_only",
    "open_autosave",
    "open_ambiguous_instance",
    "unsupported_live_update",
}

_EXPECTED_SAVE_LOCK = threading.Lock()
_EXPECTED_SAVE_DEADLINES: dict[str, float] = {}


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class WorkbookOrderCell:
    sheet: str
    row: int
    column: int
    address: str
    order_name: str
    group_key: str = ""
    candidate_index: int = 0
    header_name: str = ""
    fallback_used: bool = False
    fallback_from: str = ""


@dataclass(frozen=True)
class WorkbookRule:
    header_groups: tuple[frozenset[str], ...]
    lookup_mode: str = LOOKUP_MODE_PARTNER_REF
    row_fallback_on_not_found: bool = False
    global_search_on_not_found: bool = False
    workbook_label: str = "Workbook"
    required_header_examples: tuple[str, ...] = ()

    @property
    def headers(self) -> set[str]:
        merged: set[str] = set()
        for group in self.header_groups:
            merged.update(group)
        return merged


@dataclass(frozen=True)
class WorkbookScanResult:
    cells: list[WorkbookOrderCell]
    issue_code: str = ""
    issue_message: str = ""


@dataclass(frozen=True)
class PurchaseLinkResult:
    status: str
    source_model: str = ""
    matched_field: str = ""
    record_id: int | None = None
    record_name: str = ""
    ref_value: str = ""
    state: str = ""
    vendor: str = ""
    amount: float = 0.0
    url: str = ""
    note: str = ""


@dataclass(frozen=True)
class WorkbookProcessSummary:
    workbook_path: Path
    report_path: Path | None
    backup_path: Path | None
    total_cells: int
    unique_orders: int
    linked_count: int
    status_counts: dict[str, int]
    workbook_state: str = "closed"
    live_update_used: bool = False


@dataclass
class WorkbookAccessContext:
    status: str
    workbook_path: Path
    backend: str
    details: str = ""
    instance_pid: int | None = None
    read_only: bool = False
    autosave_on: bool = False
    application: Any = field(default=None, repr=False, compare=False)
    workbook: Any = field(default=None, repr=False, compare=False)

    @property
    def is_open(self) -> bool:
        return self.status != "closed"

    @property
    def is_live_writable(self) -> bool:
        return self.status in LIVE_UPDATE_READY_STATES and self.workbook is not None


class WorkbookAccessError(RuntimeError):
    def __init__(self, access: WorkbookAccessContext, message: str) -> None:
        super().__init__(message)
        self.access = access


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def many2one_name(value: Any) -> str:
    if isinstance(value, (list, tuple)) and len(value) >= 2:
        return str(value[1] or "")
    return str(value or "")


def normalize_header(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\xa0", " ").replace("\u00c2\u00b0", "\u00b0").replace("\u00ba", "\u00b0")
    return re.sub(r"\s+", " ", text.strip()).casefold()


def normalize_order(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_http_url(value: Any) -> bool:
    raw = str(value or "").strip()
    if not raw:
        return False
    parsed = urlparse(raw)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def normalize_odoo_url(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    parsed = urlparse(raw)
    if parsed.scheme and parsed.netloc:
        path = parsed.path.rstrip("/")
        normalized = f"{parsed.scheme}://{parsed.netloc}"
        if path:
            normalized += path
        return normalized
    return raw.rstrip("/")


def build_odoo_record_url(base_url: str, model: str, record_id: int | str) -> str:
    normalized_base = normalize_odoo_url(base_url).rstrip("/")
    normalized_model = str(model or "").strip()
    if normalized_model == "purchase.order":
        return f"{normalized_base}/odoo/purchase/{record_id}"
    query = urlencode({"id": record_id, "model": normalized_model, "view_type": "form"})
    return f"{normalized_base}/web#{query}"


def validate_odoo_settings(
    odoo_url: str,
    odoo_db: str,
    odoo_login: str,
    record_url_example: str = "",
) -> None:
    normalized_url = normalize_odoo_url(odoo_url)
    db_value = str(odoo_db or "").strip()
    login_value = str(odoo_login or "").strip()
    record_value = str(record_url_example or "").strip()
    if not normalized_url or not is_http_url(normalized_url):
        raise ValueError("Odoo URL must be a full http/https URL, for example: https://sphe.cloudoo.ma")
    parsed_odoo_url = urlparse(normalized_url)
    if parsed_odoo_url.path not in {"", "/"}:
        raise ValueError("Odoo URL must be the server root only, for example: https://sphe.cloudoo.ma")
    if not db_value:
        raise ValueError("Odoo database is required.")
    if is_http_url(db_value):
        raise ValueError("Odoo database must be the database name only, not a URL.")
    if any(char in db_value for char in ("/", "?", "#", "&", " ")):
        raise ValueError("Odoo database contains invalid URL/path characters. Use only the database name.")
    if not login_value:
        raise ValueError("Odoo login is required.")
    if record_value:
        if not is_http_url(record_value):
            raise ValueError("Purchase URL example must be a full http/https URL or left blank.")
        if urlparse(record_value).netloc.casefold() != urlparse(normalized_url).netloc.casefold():
            raise ValueError("Purchase URL example must use the same host as the Odoo URL.")


def _clean_fault_text(exc: Exception) -> str:
    if isinstance(exc, xmlrpc.client.Fault):
        return str(exc.faultString or "").strip()
    return str(exc).strip()


def explain_odoo_exception(exc: Exception, *, odoo_url: str, odoo_db: str) -> RuntimeError:
    db_value = str(odoo_db or "").strip()
    if is_http_url(db_value):
        return RuntimeError(
            "The Odoo database field contains a URL. Put only the database name in Odoo database "
            "(for example: sphe.cloudoo.ma)."
        )

    fault_text = _clean_fault_text(exc)
    lowered = fault_text.casefold()
    if "remaining connection slots are reserved" in lowered:
        return RuntimeError(
            "Odoo is temporarily overloaded: PostgreSQL connection slots are exhausted. "
            "Retry in a few minutes or contact the Odoo administrator."
        )
    if "database" in lowered and "does not exist" in lowered:
        return RuntimeError(f"Odoo database not found: '{db_value}'. Check the Odoo database field.")
    if "keyerror" in lowered and db_value and db_value.casefold() in lowered:
        return RuntimeError(f"Odoo database is invalid: '{db_value}'. Check the Odoo database field.")
    if isinstance(exc, xmlrpc.client.ProtocolError):
        return RuntimeError(
            f"Could not reach Odoo at {normalize_odoo_url(odoo_url)}. "
            "Check the Odoo URL, internet access, or server availability."
        )
    if isinstance(exc, OSError):
        return RuntimeError(
            f"Network error while contacting Odoo at {normalize_odoo_url(odoo_url)}. "
            "Check internet access and firewall settings."
        )
    if fault_text:
        first_line = fault_text.splitlines()[0].strip()
        return RuntimeError(first_line)
    return RuntimeError("Unexpected error while contacting Odoo.")


def is_transient_odoo_rpc_error(exc: Exception) -> bool:
    text = _clean_fault_text(exc).casefold()
    transient_markers = (
        "remaining connection slots are reserved",
        "too many clients",
        "connection reset",
        "connection refused",
        "connection timed out",
        "temporarily unavailable",
        "server closed the connection",
        "bad gateway",
        "gateway timeout",
        "service unavailable",
    )
    return any(marker in text for marker in transient_markers) or isinstance(exc, OSError)


def parse_amount(value: Any) -> float | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    cleaned = re.sub(r"[^\d,.\-]", "", raw).strip()
    if cleaned in {"", "-", ".", ",", "-.", "-,"}:
        return None
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "")
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _achats_local_workbook_rule() -> WorkbookRule:
    return WorkbookRule(
        header_groups=(
            frozenset(LOCAL_HEADER_VARIANTS),
            frozenset(LOCAL_COMMAND_HEADER_VARIANTS),
        ),
        lookup_mode=LOOKUP_MODE_COMMAND_REF,
        row_fallback_on_not_found=True,
        global_search_on_not_found=True,
        workbook_label="ACHATS LOCAL",
        required_header_examples=("N°FACTURE", "N commandes"),
    )


def _achats_etranger_workbook_rule() -> WorkbookRule:
    return WorkbookRule(
        header_groups=(frozenset(ETRANGER_HEADER_VARIANTS),),
        lookup_mode=LOOKUP_MODE_COMMAND_REF,
        workbook_label="ACHATS ETRANGER",
        required_header_examples=("N COMMANDE",),
    )


def _seller_previous_workbook_rule() -> WorkbookRule:
    return WorkbookRule(
        header_groups=(frozenset(ALL_HEADER_VARIANTS),),
        lookup_mode=LOOKUP_MODE_PARTNER_REF,
        workbook_label="Seller / Previous workbook",
        required_header_examples=("N commandes", "N commande", "N°FACTURE"),
    )


def workbook_rule_for_slot(slot: str, workbook_path: Path | None = None) -> WorkbookRule:
    normalized = str(slot or "").strip().casefold()
    if normalized == WORKBOOK_SLOT_ACHATS_LOCAL:
        return _achats_local_workbook_rule()
    if normalized == WORKBOOK_SLOT_ACHATS_ETRANGER:
        return _achats_etranger_workbook_rule()
    if normalized == WORKBOOK_SLOT_SELLER_PREVIOUS:
        return _seller_previous_workbook_rule()
    if workbook_path is not None:
        return workbook_rule_for_path(workbook_path)
    return _seller_previous_workbook_rule()


def workbook_rule_for_path(workbook_path: Path) -> WorkbookRule:
    workbook_name = workbook_path.name.strip().casefold()
    if workbook_name == LOCAL_WORKBOOK_FILE_NAME:
        return _achats_local_workbook_rule()
    if workbook_name == ETRANGER_WORKBOOK_FILE_NAME:
        return _achats_etranger_workbook_rule()
    return _seller_previous_workbook_rule()


def cell_address(row: int, col: int) -> str:
    letters = ""
    n = col
    while n:
        n, remainder = divmod(n - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row}"


def make_timestamp() -> str:
    return dt.datetime.now().strftime("%Y%m%d-%H%M%S")


def is_supported_workbook(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_EXTENSIONS


def normalized_workbook_key(path: Path | str) -> str:
    try:
        return str(Path(path).expanduser().resolve())
    except Exception:
        return os.path.normcase(os.path.abspath(str(path)))


_MANAGED_WORKBOOK_LOCK = threading.Lock()
_MANAGED_WORKBOOK_COUNTS: dict[str, int] = {}


def mark_expected_excel_save(path: Path, seconds: float = DEFAULT_EXCEL_SAVE_DEBOUNCE_SECONDS) -> None:
    deadline = time.monotonic() + max(float(seconds), 0.0)
    with _EXPECTED_SAVE_LOCK:
        now = time.monotonic()
        expired = [key for key, value in _EXPECTED_SAVE_DEADLINES.items() if value <= now]
        for key in expired:
            del _EXPECTED_SAVE_DEADLINES[key]
        _EXPECTED_SAVE_DEADLINES[normalized_workbook_key(path)] = deadline


def is_expected_excel_save(path: Path) -> bool:
    key = normalized_workbook_key(path)
    with _EXPECTED_SAVE_LOCK:
        deadline = _EXPECTED_SAVE_DEADLINES.get(key)
        if deadline is None:
            return False
        if deadline <= time.monotonic():
            del _EXPECTED_SAVE_DEADLINES[key]
            return False
        return True


def clear_expected_excel_save(path: Path) -> None:
    with _EXPECTED_SAVE_LOCK:
        _EXPECTED_SAVE_DEADLINES.pop(normalized_workbook_key(path), None)


@contextmanager
def managed_workbook_session(path: Path) -> Any:
    key = normalized_workbook_key(path)
    with _MANAGED_WORKBOOK_LOCK:
        _MANAGED_WORKBOOK_COUNTS[key] = _MANAGED_WORKBOOK_COUNTS.get(key, 0) + 1
    try:
        yield
    finally:
        with _MANAGED_WORKBOOK_LOCK:
            remaining = _MANAGED_WORKBOOK_COUNTS.get(key, 0) - 1
            if remaining > 0:
                _MANAGED_WORKBOOK_COUNTS[key] = remaining
            else:
                _MANAGED_WORKBOOK_COUNTS.pop(key, None)


def is_managed_workbook(path: Path) -> bool:
    with _MANAGED_WORKBOOK_LOCK:
        return _MANAGED_WORKBOOK_COUNTS.get(normalized_workbook_key(path), 0) > 0


def default_report_path(workbook_path: Path, report_dir: Path | None = None) -> Path:
    target_dir = report_dir if report_dir is not None else workbook_path.parent
    return target_dir / f"{workbook_path.stem}.purchase-link-report-{make_timestamp()}.csv"


def _set_hidden_windows_path(path: Path) -> None:
    if os.name != "nt" or not path.exists():
        return
    try:
        attrs = ctypes.windll.kernel32.GetFileAttributesW(str(path))
        if attrs in (-1, 0xFFFFFFFF):
            return
        hidden_flag = 0x02
        if not attrs & hidden_flag:
            ctypes.windll.kernel32.SetFileAttributesW(str(path), attrs | hidden_flag)
    except Exception:
        return


def prepare_backup_dir_layout(backup_root: Path) -> dict[str, Path]:
    backup_root.mkdir(parents=True, exist_ok=True)
    run_dir = backup_root / BACKUP_RUNS_DIR_NAME
    originals_dir = backup_root / BACKUP_ORIGINALS_DIR_NAME
    run_dir.mkdir(parents=True, exist_ok=True)
    originals_dir.mkdir(parents=True, exist_ok=True)
    _set_hidden_windows_path(originals_dir)

    for item in list(backup_root.iterdir()):
        if not item.is_file():
            continue
        lower_name = item.name.lower()
        destination_dir: Path | None = None
        if ".original." in lower_name:
            destination_dir = originals_dir
        elif ".backup-" in lower_name:
            destination_dir = run_dir
        if destination_dir is None:
            continue
        destination = destination_dir / item.name
        if not destination.exists():
            shutil.move(str(item), str(destination))
        if destination_dir == originals_dir:
            _set_hidden_windows_path(destination)

    return {"root": backup_root, "run_dir": run_dir, "originals_dir": originals_dir}


def build_backup_path(
    workbook_path: Path,
    backup_dir: Path | None = None,
    stable_backup_name: bool = False,
) -> Path:
    target_root = backup_dir if backup_dir is not None else workbook_path.parent
    layout = prepare_backup_dir_layout(target_root)
    target_dir = layout["originals_dir"] if stable_backup_name else layout["run_dir"]
    if stable_backup_name:
        digest = hashlib.sha1(str(workbook_path).encode("utf-8")).hexdigest()[:10]
        return target_dir / f"{workbook_path.stem}.{digest}.original{workbook_path.suffix}"
    return target_dir / f"{workbook_path.stem}.backup-{make_timestamp()}{workbook_path.suffix}"


def backup_workbook(
    workbook_path: Path,
    backup_dir: Path | None = None,
    stable_backup_name: bool = False,
    workbook: Any | None = None,
) -> Path:
    backup_path = build_backup_path(workbook_path, backup_dir, stable_backup_name=stable_backup_name)
    if stable_backup_name and backup_path.exists():
        _set_hidden_windows_path(backup_path)
        return backup_path
    if workbook is not None:
        workbook.SaveCopyAs(str(backup_path))
        if stable_backup_name:
            _set_hidden_windows_path(backup_path)
        return backup_path
    shutil.copy2(workbook_path, backup_path)
    if stable_backup_name:
        _set_hidden_windows_path(backup_path)
    return backup_path


@contextmanager
def com_scope() -> Any:
    pythoncom.CoInitialize()
    try:
        yield
    finally:
        pythoncom.CoUninitialize()


def open_excel(visible: bool) -> Any:
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = bool(visible)
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    return excel


def same_workbook_path(left: Any, right: Path) -> bool:
    try:
        return Path(str(left)).expanduser().resolve() == right.expanduser().resolve()
    except Exception:
        return os.path.normcase(os.path.abspath(str(left))) == os.path.normcase(os.path.abspath(str(right)))


def safe_excel_bool_property(obj: Any, name: str) -> bool:
    try:
        return bool(getattr(obj, name))
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Odoo Client (purchase orders)
# ---------------------------------------------------------------------------

class OdooClient:
    """Search Odoo purchase orders in Achats using workbook-specific rules."""

    def __init__(self, url: str, db: str, login: str, api_key: str) -> None:
        self.url = normalize_odoo_url(url)
        self.db = str(db or "").strip()
        self.login = str(login or "").strip()
        self.api_key = api_key
        self.uid: int | None = None
        transport = (xmlrpc.client.SafeTransport() if self.url.startswith("https")
                     else xmlrpc.client.Transport())
        transport._extra_headers = [("Connection", "keep-alive")]
        self.common = xmlrpc.client.ServerProxy(
            f"{self.url}/xmlrpc/2/common", transport=transport)
        self.models = xmlrpc.client.ServerProxy(
            f"{self.url}/xmlrpc/2/object", transport=transport)

    def authenticate(self) -> int:
        validate_odoo_settings(self.url, self.db, self.login)
        uid = None
        for attempt in range(ODOO_RPC_MAX_ATTEMPTS):
            try:
                uid = self.common.authenticate(self.db, self.login, self.api_key, {})
                break
            except Exception as exc:
                if attempt < ODOO_RPC_MAX_ATTEMPTS - 1 and is_transient_odoo_rpc_error(exc):
                    time.sleep(ODOO_RPC_RETRY_BASE_SECONDS * (2 ** attempt))
                    continue
                raise explain_odoo_exception(exc, odoo_url=self.url, odoo_db=self.db) from exc
        if not uid:
            raise RuntimeError("Odoo authentication failed. Check ODOO_DB, ODOO_LOGIN, and ODOO_API_KEY.")
        self.uid = int(uid)
        return self.uid

    def _call(self, model: str, method: str, args: list, kwargs: dict | None = None) -> Any:
        if self.uid is None:
            self.authenticate()
        for attempt in range(ODOO_RPC_MAX_ATTEMPTS):
            try:
                return self.models.execute_kw(
                    self.db, self.uid, self.api_key, model, method, args, kwargs or {})
            except Exception as exc:
                if attempt < ODOO_RPC_MAX_ATTEMPTS - 1 and is_transient_odoo_rpc_error(exc):
                    time.sleep(ODOO_RPC_RETRY_BASE_SECONDS * (2 ** attempt))
                    continue
                raise explain_odoo_exception(exc, odoo_url=self.url, odoo_db=self.db) from exc
        raise RuntimeError("Odoo RPC call failed after retries.")

    def search_purchase_orders_by_partner_ref(self, ref: str, operator: str) -> list[dict]:
        return self._call(
            "purchase.order", "search_read",
            [[("partner_ref", operator, ref)]],
            {"fields": PURCHASE_ORDER_FIELDS, "limit": 10,
             "order": "date_order desc, id desc"})

    def search_purchase_orders_by_name(self, ref: str, operator: str) -> list[dict]:
        return self._call(
            "purchase.order", "search_read",
            [[("name", operator, ref)]],
            {"fields": PURCHASE_ORDER_FIELDS, "limit": 10,
             "order": "date_order desc, id desc"})

    def search_purchase_orders_by_amount_exact(self, amount: float) -> list[dict]:
        return self._call(
            "purchase.order", "search_read",
            [[("amount_total", "=", amount)]],
            {"fields": PURCHASE_ORDER_FIELDS, "limit": 10,
             "order": "date_order desc, id desc"})

    def search_purchase_orders_by_amount_range(self, minimum: float, maximum: float) -> list[dict]:
        return self._call(
            "purchase.order", "search_read",
            [[("amount_total", ">=", minimum), ("amount_total", "<=", maximum)]],
            {"fields": PURCHASE_ORDER_FIELDS, "limit": 10,
             "order": "date_order desc, id desc"})

    def search_purchase_orders_by_exact_values(self, field_name: str, values: list[Any]) -> list[dict]:
        if field_name not in {"name", "partner_ref", "amount_total"} or not values:
            return []
        return self._call(
            "purchase.order",
            "search_read",
            [[(field_name, "in", values)]],
            {
                "fields": PURCHASE_ORDER_FIELDS,
                "limit": max(10, len(values) * 3),
                "order": "date_order desc, id desc",
            },
        )

    def resolve_purchase_ref(
        self,
        ref: str,
        lookup_mode: str = LOOKUP_MODE_PARTNER_REF,
    ) -> PurchaseLinkResult:
        term = normalize_order(ref)
        if not term:
            return PurchaseLinkResult(status="empty_ref", note="Reference was empty.")
        if lookup_mode == LOOKUP_MODE_COMMAND_REF:
            return self._resolve_by_command_reference(term)
        if lookup_mode == LOOKUP_MODE_TOTAL_AMOUNT:
            return self._resolve_by_total_amount(term)
        return self._resolve_by_partner_ref(term)

    def _resolve_by_partner_ref(self, term: str) -> PurchaseLinkResult:
        return self._resolve_by_text_field(
            term,
            field_name="partner_ref",
            search_func=self.search_purchase_orders_by_partner_ref,
            not_found_note="Reference not found in Odoo purchase orders (partner_ref).",
        )

    def _resolve_by_command_reference(self, term: str) -> PurchaseLinkResult:
        stages: list[tuple[str, str, str, Any]] = [
            ("name", "=", "exact:name", self.search_purchase_orders_by_name),
            ("partner_ref", "=", "exact:partner_ref", self.search_purchase_orders_by_partner_ref),
            ("name", "ilike", "contains:name", self.search_purchase_orders_by_name),
            ("partner_ref", "ilike", "contains:partner_ref", self.search_purchase_orders_by_partner_ref),
        ]
        for field_name, operator, default_match, search_func in stages:
            orders = search_func(term, operator)
            if not orders:
                continue
            best = self._pick_best_order(
                term,
                orders,
                field_name=field_name,
                default_match=default_match,
            )
            if best is not None:
                return best
        return PurchaseLinkResult(
            status="not_found",
            note=(
                "Reference not found in Odoo purchase orders "
                "(name / Référence commande, partner_ref / Référence fournisseur)."
            ),
        )

    def _resolve_by_total_amount(self, term: str) -> PurchaseLinkResult:
        amount = parse_amount(term)
        if amount is None:
            return PurchaseLinkResult(
                status="invalid_amount",
                note="Reference could not be parsed as amount for amount_total lookup.",
            )

        exact_orders = self.search_purchase_orders_by_amount_exact(amount)
        if exact_orders:
            best = exact_orders[0]
            return self._linked_order_result(
                best,
                matched_field="exact:amount_total",
                ref_value=str(best.get("amount_total") or ""),
            )
        for tolerance in DEFAULT_AMOUNT_TOLERANCES:
            minimum = amount - tolerance
            maximum = amount + tolerance
            range_orders = self.search_purchase_orders_by_amount_range(minimum, maximum)
            if not range_orders:
                continue
            best = self._pick_best_order_by_amount(amount, range_orders)
            return self._linked_order_result(
                best,
                matched_field=f"range:amount_total+-{tolerance:.2f}",
                ref_value=str(best.get("amount_total") or ""),
            )
        return PurchaseLinkResult(
            status="not_found",
            note="Reference not found in Odoo purchase orders (amount_total / Total).",
        )

    def _pick_best_order_by_amount(self, amount: float, orders: list[dict]) -> dict:
        def amount_value(order: dict) -> float:
            try:
                return float(order.get("amount_total") or 0.0)
            except Exception:
                return 0.0

        return sorted(
            orders,
            key=lambda order: (
                abs(amount_value(order) - amount),
                -int(order.get("id") or 0),
            ),
        )[0]

    def _resolve_by_text_field(
        self,
        term: str,
        *,
        field_name: str,
        search_func: Any,
        not_found_note: str,
    ) -> PurchaseLinkResult:
        stages: list[tuple[str, str]] = [
            ("=", f"exact:{field_name}"),
            ("ilike", f"contains:{field_name}"),
        ]
        for operator, default_match in stages:
            try:
                orders = search_func(term, operator)
            except Exception:
                orders = []
            if not orders:
                continue
            best = self._pick_best_order(
                term,
                orders,
                field_name=field_name,
                default_match=default_match,
            )
            if best is not None:
                return best
        return PurchaseLinkResult(status="not_found", note=not_found_note)

    def _pick_best_order(
        self,
        term: str,
        orders: list[dict],
        *,
        field_name: str,
        default_match: str,
    ) -> PurchaseLinkResult | None:
        folded = term.casefold()
        exact = [o for o in orders if str(o.get(field_name) or "").strip().casefold() == folded]
        matched = f"exact:{field_name}" if exact else default_match
        candidates = exact if exact else orders
        if not candidates:
            return None
        best = candidates[0]
        ref_value = str(best.get(field_name) or "")
        return self._linked_order_result(best, matched_field=matched, ref_value=ref_value)

    def _linked_order_result(
        self,
        order: dict,
        *,
        matched_field: str,
        ref_value: str,
    ) -> PurchaseLinkResult:
        best = order
        po_id = int(best["id"])
        return PurchaseLinkResult(
            status="linked",
            source_model="purchase.order",
            matched_field=matched_field,
            record_id=po_id,
            record_name=str(best.get("name") or ""),
            ref_value=ref_value,
            state=str(best.get("state") or ""),
            vendor=many2one_name(best.get("partner_id")),
            amount=float(best.get("amount_total") or 0),
            url=build_odoo_record_url(self.url, "purchase.order", po_id),
            note=f"Matched purchase order via {matched_field}.",
        )


# ---------------------------------------------------------------------------
# Excel session management
# ---------------------------------------------------------------------------

def inspect_workbook_access_state(
    workbook_path: Path,
    *,
    excel_session_backend: str = DEFAULT_EXCEL_SESSION_BACKEND,
    allow_live_update_with_autosave: bool = False,
) -> WorkbookAccessContext:
    with com_scope():
        return attach_workbook_session(
            workbook_path,
            excel_session_backend=excel_session_backend,
            allow_live_update_with_autosave=allow_live_update_with_autosave,
            attach_objects=False,
        )


def attach_workbook_session(
    workbook_path: Path,
    *,
    excel_session_backend: str = DEFAULT_EXCEL_SESSION_BACKEND,
    allow_live_update_with_autosave: bool = False,
    attach_objects: bool = True,
) -> WorkbookAccessContext:
    workbook_path = workbook_path.expanduser().resolve()
    backend = str(excel_session_backend or DEFAULT_EXCEL_SESSION_BACKEND).strip().casefold()
    if backend not in {"xlwings", "pywin32"}:
        backend = DEFAULT_EXCEL_SESSION_BACKEND
    attempted: list[WorkbookAccessContext] = []
    order = [backend, "pywin32" if backend == "xlwings" else "xlwings"]
    for candidate_backend in order:
        if candidate_backend == "xlwings":
            access = _attach_xlwings(workbook_path,
                                     allow_live_update_with_autosave=allow_live_update_with_autosave,
                                     attach_objects=attach_objects)
        else:
            access = _attach_pywin32(workbook_path,
                                     allow_live_update_with_autosave=allow_live_update_with_autosave,
                                     attach_objects=attach_objects)
        attempted.append(access)
        if access.status != "closed":
            return access
    for access in attempted:
        if access.status == "unsupported_live_update":
            return access
    return attempted[0]


def _build_live_access(
    workbook_path: Path, *, backend: str, workbook: Any, application: Any,
    instance_pid: int | None, allow_live_update_with_autosave: bool,
    attach_objects: bool,
) -> WorkbookAccessContext:
    read_only = safe_excel_bool_property(workbook, "ReadOnly")
    autosave_on = safe_excel_bool_property(workbook, "AutoSaveOn")
    if read_only:
        status, details = "open_read_only", "Workbook is open in Excel as read-only."
    elif autosave_on and not allow_live_update_with_autosave:
        status, details = "open_autosave", "Workbook is open with AutoSave enabled."
    else:
        status, details = "open_writable", "Workbook is open and can be updated live."
    return WorkbookAccessContext(
        status=status, workbook_path=workbook_path, backend=backend,
        details=details, instance_pid=instance_pid,
        read_only=read_only, autosave_on=autosave_on,
        application=application if attach_objects else None,
        workbook=workbook if attach_objects else None,
    )


def _attach_xlwings(
    workbook_path: Path, *, allow_live_update_with_autosave: bool,
    attach_objects: bool,
) -> WorkbookAccessContext:
    if xw is None:
        return WorkbookAccessContext(
            status="unsupported_live_update", workbook_path=workbook_path,
            backend="xlwings", details="xlwings is not installed.")
    try:
        pids = list(xw.apps.keys())
    except Exception as exc:
        return WorkbookAccessContext(
            status="unsupported_live_update", workbook_path=workbook_path,
            backend="xlwings", details=f"xlwings error: {exc}")
    matches: list[tuple[int, Any, Any]] = []
    for pid in pids:
        try:
            app = xw.apps[pid]
            books = list(app.books)
        except Exception:
            continue
        for book in books:
            fullname = ""
            try:
                fullname = str(book.fullname or "")
            except Exception:
                pass
            if not fullname:
                try:
                    fullname = str(book.api.FullName or "")
                except Exception:
                    pass
            if fullname and same_workbook_path(fullname, workbook_path):
                matches.append((int(pid), app, book))
    if not matches:
        return WorkbookAccessContext(status="closed", workbook_path=workbook_path, backend="xlwings")
    if len(matches) > 1:
        return WorkbookAccessContext(
            status="open_ambiguous_instance", workbook_path=workbook_path,
            backend="xlwings", details="Open in more than one Excel instance.")
    pid, app, book = matches[0]
    return _build_live_access(
        workbook_path, backend="xlwings", workbook=book.api, application=app.api,
        instance_pid=pid, allow_live_update_with_autosave=allow_live_update_with_autosave,
        attach_objects=attach_objects)


def _attach_pywin32(
    workbook_path: Path, *, allow_live_update_with_autosave: bool,
    attach_objects: bool,
) -> WorkbookAccessContext:
    try:
        application = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        return WorkbookAccessContext(status="closed", workbook_path=workbook_path, backend="pywin32")
    matches: list[Any] = []
    for candidate in application.Workbooks:
        fullname = ""
        try:
            fullname = str(candidate.FullName or "")
        except Exception:
            pass
        if fullname and same_workbook_path(fullname, workbook_path):
            matches.append(candidate)
    if not matches:
        return WorkbookAccessContext(status="closed", workbook_path=workbook_path, backend="pywin32")
    if len(matches) > 1:
        return WorkbookAccessContext(
            status="open_ambiguous_instance", workbook_path=workbook_path,
            backend="pywin32", details="Open more than once in active Excel.")
    return _build_live_access(
        workbook_path, backend="pywin32", workbook=matches[0], application=application,
        instance_pid=None, allow_live_update_with_autosave=allow_live_update_with_autosave,
        attach_objects=attach_objects)


# ---------------------------------------------------------------------------
# Workbook scanning — auto-detects headers based on workbook rule
# ---------------------------------------------------------------------------

def _find_purchase_headers(
    all_values: tuple,
    row_count: int,
    col_count: int,
    header_groups: tuple[frozenset[str], ...],
) -> list[tuple[int, int, int, str]]:
    headers: list[tuple[int, int, int, str]] = []
    seen: set[tuple[int, int, int]] = set()
    for ri in range(min(row_count, 80)):
        for ci in range(col_count):
            norm = normalize_header(all_values[ri][ci])
            if not norm:
                continue
            for group_index, header_group in enumerate(header_groups):
                if norm in header_group:
                    key = (group_index, ri, ci)
                    if key not in seen:
                        seen.add(key)
                        headers.append((group_index, ri, ci, norm))
    return headers


def collect_workbook_orders(
    workbook: Any,
    workbook_rule: WorkbookRule,
) -> WorkbookScanResult:
    cells: list[WorkbookOrderCell] = []
    found_header_groups: set[int] = set()
    for sheet in workbook.Worksheets:
        sheet_name = sheet.Name
        used_range = sheet.UsedRange
        all_values = used_range.Value
        if all_values is None:
            continue
        if not isinstance(all_values, tuple):
            all_values = ((all_values,),)
        row_count = len(all_values)
        if row_count == 0:
            continue
        if not isinstance(all_values[0], tuple):
            all_values = tuple((v,) for v in all_values)
            row_count = len(all_values)
        col_count = len(all_values[0])
        first_row = int(used_range.Row)
        first_col = int(used_range.Column)

        sheet_cells, sheet_groups = _scan_sheet_values(
            sheet_name,
            all_values,
            first_row,
            first_col,
            workbook_rule,
        )
        cells.extend(sheet_cells)
        found_header_groups.update(sheet_groups)
    return _finalize_scan_result(cells, found_header_groups, workbook_rule)


def _scan_sheet_values(
    sheet_name: str,
    all_values: tuple[tuple[Any, ...], ...] | tuple,
    first_row: int,
    first_col: int,
    workbook_rule: WorkbookRule,
) -> tuple[list[WorkbookOrderCell], set[int]]:
    row_count = len(all_values)
    if row_count == 0:
        return [], set()
    col_count = len(all_values[0])
    headers = _find_purchase_headers(all_values, row_count, col_count, workbook_rule.header_groups)
    if not headers:
        return [], set()
    found_header_groups = {group_index for group_index, _, _, _ in headers}
    if workbook_rule.row_fallback_on_not_found:
        return (
            _collect_row_fallback_cells(
                sheet_name,
                all_values,
                first_row,
                first_col,
                headers,
            ),
            found_header_groups,
        )
    return (
        _collect_cells_by_headers(
            sheet_name,
            all_values,
            first_row,
            first_col,
            headers,
        ),
        found_header_groups,
    )


def _collect_cells_by_headers(
    sheet_name: str,
    all_values: tuple[tuple[Any, ...], ...] | tuple,
    first_row: int,
    first_col: int,
    headers: list[tuple[int, int, int, str]],
) -> list[WorkbookOrderCell]:
    cells: list[WorkbookOrderCell] = []
    seen_cells: set[tuple[str, int, int]] = set()
    for _, header_ri, header_ci, header_name in headers:
        for ri in range(header_ri + 1, len(all_values)):
            val = normalize_order(all_values[ri][header_ci])
            if not _is_candidate_value(val):
                continue
            actual_row = first_row + ri
            actual_col = first_col + header_ci
            cell_key = (sheet_name, actual_row, actual_col)
            if cell_key in seen_cells:
                continue
            seen_cells.add(cell_key)
            address = cell_address(actual_row, actual_col)
            cells.append(
                WorkbookOrderCell(
                    sheet=sheet_name,
                    row=actual_row,
                    column=actual_col,
                    address=address,
                    order_name=val,
                    group_key=f"{sheet_name}!{address}",
                    header_name=header_name,
                )
            )
    return cells


def _collect_row_fallback_cells(
    sheet_name: str,
    all_values: tuple[tuple[Any, ...], ...] | tuple,
    first_row: int,
    first_col: int,
    headers: list[tuple[int, int, int, str]],
) -> list[WorkbookOrderCell]:
    cells_by_row_and_group: dict[tuple[str, int, int], WorkbookOrderCell] = {}
    for group_index, header_ri, header_ci, header_name in sorted(headers, key=lambda item: (item[0], item[1], item[2])):
        for ri in range(header_ri + 1, len(all_values)):
            val = normalize_order(all_values[ri][header_ci])
            if not _is_candidate_value(val):
                continue
            actual_row = first_row + ri
            row_group_key = (sheet_name, actual_row, group_index)
            if row_group_key in cells_by_row_and_group:
                continue
            actual_col = first_col + header_ci
            address = cell_address(actual_row, actual_col)
            cells_by_row_and_group[row_group_key] = WorkbookOrderCell(
                sheet=sheet_name,
                row=actual_row,
                column=actual_col,
                address=address,
                order_name=val,
                group_key=f"{sheet_name}!row:{actual_row}",
                candidate_index=group_index,
                header_name=header_name,
            )
    ordered = sorted(
        cells_by_row_and_group.values(),
        key=lambda cell: (cell.sheet.casefold(), cell.row, cell.candidate_index, cell.column),
    )
    return ordered


def _is_candidate_value(value: str) -> bool:
    if not value:
        return False
    if value.strip().startswith("="):
        return False
    return value.casefold() not in {"none", "nan"}


def _finalize_scan_result(
    cells: list[WorkbookOrderCell],
    found_header_groups: set[int],
    workbook_rule: WorkbookRule,
) -> WorkbookScanResult:
    if found_header_groups:
        return WorkbookScanResult(cells=cells)
    examples = ", ".join(workbook_rule.required_header_examples or tuple(sorted(workbook_rule.headers)))
    message = f"{workbook_rule.workbook_label} is missing the required header. Expected one of: {examples}."
    return WorkbookScanResult(cells=[], issue_code="missing_required_header", issue_message=message)


def scan_workbook_orders_from_file(
    workbook_path: Path,
    visible_excel: bool,
    workbook_rule: WorkbookRule,
    allow_com_fallback: bool = True,
) -> WorkbookScanResult:
    suffix = workbook_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"} and load_workbook is not None:
        try:
            return scan_workbook_orders_with_openpyxl(workbook_path, workbook_rule)
        except Exception:
            if not allow_com_fallback:
                raise
    return scan_workbook_orders_with_excel_com(workbook_path, visible_excel, workbook_rule)


def scan_workbook_orders_with_openpyxl(
    workbook_path: Path,
    workbook_rule: WorkbookRule,
) -> WorkbookScanResult:
    workbook = load_workbook(
        filename=str(workbook_path),
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        cells: list[WorkbookOrderCell] = []
        found_header_groups: set[int] = set()
        for sheet in workbook.worksheets:
            sheet_cells, sheet_groups = _scan_openpyxl_sheet(sheet, workbook_rule)
            cells.extend(sheet_cells)
            found_header_groups.update(sheet_groups)
        return _finalize_scan_result(cells, found_header_groups, workbook_rule)
    finally:
        workbook.close()


def _scan_openpyxl_sheet(sheet: Any, workbook_rule: WorkbookRule) -> tuple[list[WorkbookOrderCell], set[int]]:
    headers: list[tuple[int, int, int, str]] = []
    seen: set[tuple[int, int, int]] = set()
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
        for column_index, value in enumerate(row, start=1):
            norm = normalize_header(value)
            if not norm:
                continue
            for group_index, header_group in enumerate(workbook_rule.header_groups):
                if norm in header_group:
                    key = (group_index, row_index, column_index)
                    if key not in seen:
                        seen.add(key)
                        headers.append((group_index, row_index, column_index, norm))
    if not headers:
        return [], set()

    found_header_groups = {group_index for group_index, _, _, _ in headers}
    min_data_row = min(row_index for _, row_index, _, _ in headers) + 1
    if workbook_rule.row_fallback_on_not_found:
        return (
            _collect_openpyxl_row_fallback_cells(sheet.title, sheet, headers, min_data_row),
            found_header_groups,
        )
    return (
        _collect_openpyxl_cells_by_headers(sheet.title, sheet, headers, min_data_row),
        found_header_groups,
    )


def _value_at_openpyxl_column(row: tuple[Any, ...], column_index: int) -> Any:
    offset = column_index - 1
    if offset < 0 or offset >= len(row):
        return None
    return row[offset]


def _collect_openpyxl_cells_by_headers(
    sheet_name: str,
    sheet: Any,
    headers: list[tuple[int, int, int, str]],
    min_data_row: int,
) -> list[WorkbookOrderCell]:
    cells: list[WorkbookOrderCell] = []
    seen_cells: set[tuple[str, int, int]] = set()
    for row_index, row in enumerate(sheet.iter_rows(min_row=min_data_row, values_only=True), start=min_data_row):
        for _, header_row, header_column, header_name in headers:
            if row_index <= header_row:
                continue
            val = normalize_order(_value_at_openpyxl_column(row, header_column))
            if not _is_candidate_value(val):
                continue
            cell_key = (sheet_name, row_index, header_column)
            if cell_key in seen_cells:
                continue
            seen_cells.add(cell_key)
            address = cell_address(row_index, header_column)
            cells.append(
                WorkbookOrderCell(
                    sheet=sheet_name,
                    row=row_index,
                    column=header_column,
                    address=address,
                    order_name=val,
                    group_key=f"{sheet_name}!{address}",
                    header_name=header_name,
                )
            )
    return cells


def _collect_openpyxl_row_fallback_cells(
    sheet_name: str,
    sheet: Any,
    headers: list[tuple[int, int, int, str]],
    min_data_row: int,
) -> list[WorkbookOrderCell]:
    cells_by_row_and_group: dict[tuple[str, int, int], WorkbookOrderCell] = {}
    ordered_headers = sorted(headers, key=lambda item: (item[0], item[1], item[2]))
    for row_index, row in enumerate(sheet.iter_rows(min_row=min_data_row, values_only=True), start=min_data_row):
        for group_index, header_row, header_column, header_name in ordered_headers:
            if row_index <= header_row:
                continue
            val = normalize_order(_value_at_openpyxl_column(row, header_column))
            if not _is_candidate_value(val):
                continue
            row_group_key = (sheet_name, row_index, group_index)
            if row_group_key in cells_by_row_and_group:
                continue
            address = cell_address(row_index, header_column)
            cells_by_row_and_group[row_group_key] = WorkbookOrderCell(
                sheet=sheet_name,
                row=row_index,
                column=header_column,
                address=address,
                order_name=val,
                group_key=f"{sheet_name}!row:{row_index}",
                candidate_index=group_index,
                header_name=header_name,
            )
    return sorted(
        cells_by_row_and_group.values(),
        key=lambda cell: (cell.sheet.casefold(), cell.row, cell.candidate_index, cell.column),
    )


def scan_workbook_orders_with_excel_com(
    workbook_path: Path,
    visible_excel: bool,
    workbook_rule: WorkbookRule,
) -> WorkbookScanResult:
    excel = open_excel(visible_excel)
    wb = None
    try:
        with managed_workbook_session(workbook_path):
            wb = excel.Workbooks.Open(
                str(workbook_path),
                ReadOnly=True,
                UpdateLinks=0,
                IgnoreReadOnlyRecommended=True,
            )
            return collect_workbook_orders(wb, workbook_rule)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()


def scan_workbook_orders(
    workbook_path: Path, visible_excel: bool,
    access: WorkbookAccessContext | None = None,
    workbook_rule: WorkbookRule | None = None,
) -> WorkbookScanResult:
    workbook_rule = workbook_rule or workbook_rule_for_path(workbook_path)
    if access is not None and access.workbook is not None and access.is_open:
        return collect_workbook_orders(access.workbook, workbook_rule)
    with com_scope():
        return scan_workbook_orders_from_file(
            workbook_path,
            visible_excel,
            workbook_rule,
        )


def select_cells_for_results(
    cells: list[WorkbookOrderCell],
    results: dict[str, PurchaseLinkResult],
    workbook_rule: WorkbookRule,
) -> list[WorkbookOrderCell]:
    if not workbook_rule.row_fallback_on_not_found:
        return cells
    grouped: dict[str, list[WorkbookOrderCell]] = {}
    group_order: list[str] = []
    for cell in cells:
        if cell.group_key not in grouped:
            grouped[cell.group_key] = []
            group_order.append(cell.group_key)
        grouped[cell.group_key].append(cell)

    selected: list[WorkbookOrderCell] = []
    for group_key in group_order:
        ordered = sorted(grouped[group_key], key=lambda cell: (cell.candidate_index, cell.column))
        primary = ordered[0]
        chosen = primary
        primary_result = results.get(primary.order_name, PurchaseLinkResult(status="unknown"))
        if primary_result.status == "not_found":
            for fallback in ordered[1:]:
                fallback_result = results.get(fallback.order_name, PurchaseLinkResult(status="unknown"))
                if fallback_result.status != "not_found":
                    chosen = replace(
                        fallback,
                        fallback_used=True,
                        fallback_from=primary.order_name,
                    )
                    break
        selected.append(chosen)
    return selected


def selected_status_counts(
    cells: list[WorkbookOrderCell],
    results: dict[str, PurchaseLinkResult],
    scan_result: WorkbookScanResult,
) -> dict[str, int]:
    if scan_result.issue_code and not cells:
        return {scan_result.issue_code: 1}
    counts: dict[str, int] = {}
    for cell in cells:
        result = results.get(cell.order_name, PurchaseLinkResult(status="unknown"))
        counts[result.status] = counts.get(result.status, 0) + 1
    fallback_used = sum(1 for cell in cells if cell.fallback_used)
    if fallback_used:
        counts["row_fallback_used"] = fallback_used
    return counts


def linked_cell_count(cells: list[WorkbookOrderCell], results: dict[str, PurchaseLinkResult]) -> int:
    return sum(
        1
        for cell in cells
        if (results.get(cell.order_name) or PurchaseLinkResult(status="unknown")).status == "linked"
    )


# ---------------------------------------------------------------------------
# Hyperlink writing
# ---------------------------------------------------------------------------

def save_workbook_with_suppressed_events(workbook: Any, save_debounce_seconds: float) -> None:
    workbook_path = Path(str(workbook.FullName)).expanduser().resolve()
    application = workbook.Application
    previous_enable_events = None
    try:
        previous_enable_events = bool(application.EnableEvents)
        application.EnableEvents = False
    except Exception:
        previous_enable_events = None
    mark_expected_excel_save(workbook_path, save_debounce_seconds)
    try:
        workbook.Save()
    except Exception:
        clear_expected_excel_save(workbook_path)
        raise
    finally:
        if previous_enable_events is not None:
            try:
                application.EnableEvents = previous_enable_events
            except Exception:
                pass


def apply_links_to_workbook(
    workbook: Any,
    cells: list[WorkbookOrderCell],
    results: dict[str, PurchaseLinkResult],
) -> int:
    application = workbook.Application
    prev_screen = prev_calc = prev_events = None
    try:
        try:
            prev_screen = application.ScreenUpdating
            prev_calc = application.Calculation
            prev_events = application.EnableEvents
            application.ScreenUpdating = False
            application.Calculation = -4135
            application.EnableEvents = False
        except Exception:
            pass
        count = 0
        sheet_map = {s.Name: s for s in workbook.Worksheets}
        for cell_info in cells:
            result = results.get(cell_info.order_name)
            if result is None or result.status != "linked" or not result.url:
                continue
            sheet = sheet_map.get(cell_info.sheet)
            if sheet is None:
                continue
            cell = sheet.Cells(cell_info.row, cell_info.column)
            try:
                cell.Hyperlinks.Delete()
            except Exception:
                pass
            tip = f"Open {result.source_model} in Odoo"
            sheet.Hyperlinks.Add(
                Anchor=cell, Address=result.url,
                TextToDisplay=cell_info.order_name, ScreenTip=tip)
            count += 1
        return count
    finally:
        try:
            if prev_events is not None:
                application.EnableEvents = prev_events
            if prev_calc is not None:
                application.Calculation = prev_calc
            if prev_screen is not None:
                application.ScreenUpdating = prev_screen
        except Exception:
            pass


def write_links_with_excel(
    workbook_path: Path, cells: list[WorkbookOrderCell],
    results: dict[str, PurchaseLinkResult],
    visible_excel: bool, save_debounce_seconds: float,
) -> int:
    excel = open_excel(visible_excel)
    wb = None
    try:
        with managed_workbook_session(workbook_path):
            wb = excel.Workbooks.Open(str(workbook_path), ReadOnly=False,
                                      UpdateLinks=0, IgnoreReadOnlyRecommended=True)
            if bool(wb.ReadOnly):
                raise RuntimeError("Excel opened the workbook as read-only.")
            count = apply_links_to_workbook(wb, cells, results)
            if count > 0:
                save_workbook_with_suppressed_events(wb, save_debounce_seconds)
            return count
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()


def write_links_with_openpyxl(
    workbook_path: Path,
    cells: list[WorkbookOrderCell],
    results: dict[str, PurchaseLinkResult],
) -> int:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to update .xlsx/.xlsm files without Excel.")
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(
        filename=str(workbook_path),
        read_only=False,
        keep_vba=keep_vba,
        data_only=False,
        keep_links=True,
    )
    try:
        sheet_map = {sheet.title: sheet for sheet in workbook.worksheets}
        count = 0
        for cell_info in cells:
            result = results.get(cell_info.order_name)
            if result is None or result.status != "linked" or not result.url:
                continue
            sheet = sheet_map.get(cell_info.sheet)
            if sheet is None:
                continue
            cell = sheet.cell(row=cell_info.row, column=cell_info.column)
            cell.hyperlink = None
            cell.hyperlink = result.url
            if cell.value is None:
                cell.value = cell_info.order_name
            count += 1
        if count:
            workbook.save(str(workbook_path))
        return count
    finally:
        workbook.close()


def _ooxml_qname(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def _ooxml_workbook_sheet_map(zip_file: zipfile.ZipFile) -> dict[str, str]:
    workbook_root = ET.fromstring(zip_file.read("xl/workbook.xml"))
    rels_root = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))
    rels: dict[str, str] = {}
    for rel in rels_root.findall(_ooxml_qname(OOXML_PACKAGE_REL_NS, "Relationship")):
        rel_id = str(rel.attrib.get("Id") or "")
        target = str(rel.attrib.get("Target") or "")
        if rel_id and target:
            if target.startswith("/"):
                rels[rel_id] = target.lstrip("/")
            else:
                rels[rel_id] = posixpath.normpath(posixpath.join("xl", target))

    sheet_map: dict[str, str] = {}
    for sheet in workbook_root.findall(f".//{_ooxml_qname(OOXML_MAIN_NS, 'sheet')}"):
        sheet_name = str(sheet.attrib.get("name") or "")
        rel_id = str(sheet.attrib.get(_ooxml_qname(OOXML_DOC_REL_NS, "id")) or "")
        sheet_path = rels.get(rel_id)
        if sheet_name and sheet_path:
            sheet_map[sheet_name] = sheet_path
    return sheet_map


def _ooxml_sheet_rels_path(sheet_xml_path: str) -> str:
    folder, filename = posixpath.split(sheet_xml_path)
    return posixpath.join(folder, "_rels", f"{filename}.rels")


def _ooxml_xml_bytes(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _ooxml_next_relationship_id(rels_root: ET.Element) -> str:
    max_id = 0
    for rel in rels_root.findall(_ooxml_qname(OOXML_PACKAGE_REL_NS, "Relationship")):
        rel_id = str(rel.attrib.get("Id") or "")
        match = re.fullmatch(r"rId(\d+)", rel_id)
        if match:
            max_id = max(max_id, int(match.group(1)))
    return f"rId{max_id + 1}"


def _ooxml_root_tag_span(xml_text: str, root_name: str) -> tuple[int, int]:
    match = re.search(rf"<(?:[A-Za-z_][\w.-]*:)?{re.escape(root_name)}\b[^>]*>", xml_text)
    if not match:
        raise RuntimeError(f"OOXML part is missing the {root_name} root element.")
    return match.span()


def _worksheet_root_tag_span(sheet_text: str) -> tuple[int, int]:
    return _ooxml_root_tag_span(sheet_text, "worksheet")


def _ooxml_add_namespace_to_root(xml_text: str, root_name: str, prefix: str, uri: str) -> str:
    start, end = _ooxml_root_tag_span(xml_text, root_name)
    root_tag = xml_text[start:end]
    if re.search(rf"\sxmlns:{re.escape(prefix)}=", root_tag):
        return xml_text
    updated_tag = root_tag[:-1] + f" xmlns:{prefix}={quoteattr(uri)}>"
    return xml_text[:start] + updated_tag + xml_text[end:]


def _ooxml_add_root_namespace(sheet_text: str, prefix: str, uri: str) -> str:
    return _ooxml_add_namespace_to_root(sheet_text, "worksheet", prefix, uri)


def _ooxml_relationship_prefix(sheet_text: str) -> tuple[str, str]:
    start, end = _worksheet_root_tag_span(sheet_text)
    root_tag = sheet_text[start:end]
    match = re.search(
        rf"\sxmlns:([A-Za-z_][\w.-]*)={quoteattr(OOXML_DOC_REL_NS)}",
        root_tag,
    )
    if match:
        return match.group(1), sheet_text
    prefix = "r"
    return prefix, _ooxml_add_root_namespace(sheet_text, prefix, OOXML_DOC_REL_NS)


def _ooxml_repair_ignorable_namespace_declarations(sheet_text: str) -> str:
    return _ooxml_repair_ignorable_namespaces(sheet_text, "worksheet")


def _ooxml_repair_ignorable_namespaces(xml_text: str, root_name: str) -> str:
    start, end = _ooxml_root_tag_span(xml_text, root_name)
    root_tag = xml_text[start:end]
    match = re.search(r"\s(?:[A-Za-z_][\w.-]*:)?Ignorable=(['\"])(.*?)\1", root_tag)
    if not match:
        return xml_text
    for prefix in match.group(2).split():
        uri = OOXML_KNOWN_IGNORABLE_NAMESPACES.get(prefix)
        if uri and not re.search(rf"\sxmlns:{re.escape(prefix)}=", root_tag):
            xml_text = _ooxml_add_namespace_to_root(xml_text, root_name, prefix, uri)
            start, end = _ooxml_root_tag_span(xml_text, root_name)
            root_tag = xml_text[start:end]
    return xml_text


def _ooxml_xml_attr(element_text: str, attr_name: str) -> str:
    pattern = rf"\s{re.escape(attr_name)}=(['\"])(.*?)\1"
    match = re.search(pattern, element_text)
    return match.group(2) if match else ""


def _ooxml_xml_local_attr(element_text: str, attr_local_name: str) -> str:
    pattern = rf"\s(?:[A-Za-z_][\w.-]*:)?{re.escape(attr_local_name)}=(['\"])(.*?)\1"
    match = re.search(pattern, element_text)
    return match.group(2) if match else ""


def _ooxml_hyperlink_block_span(sheet_text: str) -> tuple[int, int] | None:
    match = re.search(r"<hyperlinks\b[^>]*>.*?</hyperlinks>", sheet_text, flags=re.DOTALL)
    return match.span() if match else None


def _ooxml_remove_sheet_hyperlinks(sheet_text: str, addresses: set[str]) -> tuple[str, set[str]]:
    block_span = _ooxml_hyperlink_block_span(sheet_text)
    if block_span is None:
        return sheet_text, set()
    start, end = block_span
    block = sheet_text[start:end]
    open_match = re.match(r"<hyperlinks\b[^>]*>", block, flags=re.DOTALL)
    close_match = re.search(r"</hyperlinks>\s*$", block, flags=re.DOTALL)
    if open_match is None or close_match is None:
        return sheet_text, set()

    inner = block[open_match.end():close_match.start()]
    removed_rel_ids: set[str] = set()

    def replace_link(match: re.Match[str]) -> str:
        element = match.group(0)
        ref = _ooxml_xml_attr(element, "ref")
        if ref not in addresses:
            return element
        rel_id = _ooxml_xml_local_attr(element, "id")
        if rel_id:
            removed_rel_ids.add(rel_id)
        return ""

    updated_inner = re.sub(
        r"<hyperlink\b[^>]*(?:/>|>.*?</hyperlink>)",
        replace_link,
        inner,
        flags=re.DOTALL,
    )
    updated_block = block[:open_match.end()] + updated_inner + block[close_match.start():]
    return sheet_text[:start] + updated_block + sheet_text[end:], removed_rel_ids


def _ooxml_hyperlinks_insert_position(sheet_text: str) -> int:
    block_span = _ooxml_hyperlink_block_span(sheet_text)
    if block_span is not None:
        return block_span[1] - len("</hyperlinks>")
    candidates: list[int] = []
    for tag in OOXML_HYPERLINK_INSERT_BEFORE_TAGS:
        match = re.search(rf"<(?:[A-Za-z_][\w.-]*:)?{tag}\b", sheet_text)
        if match:
            candidates.append(match.start())
    if candidates:
        return min(candidates)
    close_match = re.search(r"</worksheet>\s*$", sheet_text)
    if close_match:
        return close_match.start()
    raise RuntimeError("Worksheet XML is missing the closing worksheet element.")


def _ooxml_add_sheet_hyperlinks(
    sheet_text: str,
    link_rel_ids_by_address: dict[str, str],
    rel_prefix: str,
) -> str:
    if not link_rel_ids_by_address:
        return sheet_text
    block_span = _ooxml_hyperlink_block_span(sheet_text)
    if block_span is None:
        insert_at = _ooxml_hyperlinks_insert_position(sheet_text)
        sheet_text = sheet_text[:insert_at] + "<hyperlinks></hyperlinks>" + sheet_text[insert_at:]
    insert_at = _ooxml_hyperlinks_insert_position(sheet_text)
    link_xml = "".join(
        f"<hyperlink ref={quoteattr(address)} {rel_prefix}:id={quoteattr(rel_id)}/>"
        for address, rel_id in sorted(link_rel_ids_by_address.items(), key=lambda item: item[0])
    )
    return sheet_text[:insert_at] + link_xml + sheet_text[insert_at:]


def _ooxml_cell_elements(sheet_text: str) -> Any:
    return re.finditer(r"<c\b[^>]*(?:/>|>.*?</c>)", sheet_text, flags=re.DOTALL)


def _ooxml_collect_cell_style_ids(sheet_text: str, addresses: set[str]) -> dict[str, int]:
    style_ids: dict[str, int] = {}
    for address in sorted(addresses):
        match = re.search(
            rf"<c\b(?=[^>]*\br={quoteattr(address)})[^>]*(?:/>|>.*?</c>)",
            sheet_text,
            flags=re.DOTALL,
        )
        if match is None:
            continue
        cell_xml = match.group(0)
        try:
            style_ids[address] = int(_ooxml_xml_attr(cell_xml, "s") or "0")
        except ValueError:
            style_ids[address] = 0
    return style_ids


def _ooxml_set_cell_style_ids(sheet_text: str, style_by_address: dict[str, int]) -> str:
    if not style_by_address:
        return sheet_text

    def update_cell_xml(cell_xml: str, style_id: int) -> str:
        style_id_str = str(style_id)
        if re.search(r"\ss=(['\"])(.*?)\1", cell_xml):
            return re.sub(r"\ss=(['\"])(.*?)\1", f' s="{style_id_str}"', cell_xml, count=1)
        insert_at = cell_xml.find("/>")
        if insert_at < 0:
            insert_at = cell_xml.find(">")
        if insert_at < 0:
            return cell_xml
        return cell_xml[:insert_at] + f' s="{style_id_str}"' + cell_xml[insert_at:]

    updated_text = sheet_text
    for address, style_id in style_by_address.items():
        cell_pattern = re.compile(
            rf"<c\b(?=[^>]*\br={quoteattr(address)})[^>]*(?:/>|>.*?</c>)",
            flags=re.DOTALL,
        )

        def replace_target(match: re.Match[str]) -> str:
            return update_cell_xml(match.group(0), style_id)

        updated_text = cell_pattern.sub(replace_target, updated_text, count=1)
    return updated_text


def _ooxml_child_elements(parent: ET.Element, local_name: str) -> list[ET.Element]:
    return list(parent.findall(_ooxml_qname(OOXML_MAIN_NS, local_name)))


def _ooxml_ensure_styles_child(root: ET.Element, local_name: str) -> ET.Element:
    child = root.find(_ooxml_qname(OOXML_MAIN_NS, local_name))
    if child is not None:
        return child
    child = ET.Element(_ooxml_qname(OOXML_MAIN_NS, local_name), {"count": "0"})
    root.append(child)
    return child


def _ooxml_update_count(element: ET.Element) -> None:
    element.attrib["count"] = str(len(list(element)))


def _ooxml_ensure_hyperlink_font(fonts: ET.Element, base_font: ET.Element | None) -> int:
    desired = copy.deepcopy(base_font) if base_font is not None else ET.Element(_ooxml_qname(OOXML_MAIN_NS, "font"))
    color = desired.find(_ooxml_qname(OOXML_MAIN_NS, "color"))
    if color is None:
        color = ET.Element(_ooxml_qname(OOXML_MAIN_NS, "color"))
        desired.append(color)
    color.attrib.clear()
    color.attrib["rgb"] = "FF0563C1"
    if desired.find(_ooxml_qname(OOXML_MAIN_NS, "u")) is None:
        desired.append(ET.Element(_ooxml_qname(OOXML_MAIN_NS, "u")))

    desired_xml = ET.tostring(desired, encoding="utf-8")
    for index, font in enumerate(list(fonts)):
        if ET.tostring(font, encoding="utf-8") == desired_xml:
            return index
    fonts.append(desired)
    _ooxml_update_count(fonts)
    return len(list(fonts)) - 1


def _ooxml_ensure_hyperlink_xf(cell_xfs: ET.Element, base_xf: ET.Element | None, font_id: int) -> int:
    desired = copy.deepcopy(base_xf) if base_xf is not None else ET.Element(_ooxml_qname(OOXML_MAIN_NS, "xf"))
    desired.attrib["fontId"] = str(font_id)
    desired.attrib["applyFont"] = "1"
    desired_xml = ET.tostring(desired, encoding="utf-8")
    for index, xf in enumerate(list(cell_xfs)):
        if ET.tostring(xf, encoding="utf-8") == desired_xml:
            return index
    cell_xfs.append(desired)
    _ooxml_update_count(cell_xfs)
    return len(list(cell_xfs)) - 1


def _ooxml_patch_styles_for_hyperlinks(
    styles_xml: bytes,
    base_style_ids: set[int],
) -> tuple[bytes, dict[int, int]]:
    if not base_style_ids:
        return styles_xml, {}
    root = ET.fromstring(styles_xml)
    fonts = _ooxml_ensure_styles_child(root, "fonts")
    cell_xfs = _ooxml_ensure_styles_child(root, "cellXfs")
    font_list = _ooxml_child_elements(fonts, "font")
    xf_list = _ooxml_child_elements(cell_xfs, "xf")
    if not xf_list:
        cell_xfs.append(ET.Element(_ooxml_qname(OOXML_MAIN_NS, "xf"), {"fontId": "0", "fillId": "0", "borderId": "0", "numFmtId": "0"}))
        _ooxml_update_count(cell_xfs)
        xf_list = _ooxml_child_elements(cell_xfs, "xf")

    style_map: dict[int, int] = {}
    for base_style_id in sorted(base_style_ids):
        base_xf = xf_list[base_style_id] if 0 <= base_style_id < len(xf_list) else xf_list[0]
        try:
            base_font_id = int(base_xf.attrib.get("fontId", "0"))
        except ValueError:
            base_font_id = 0
        base_font = font_list[base_font_id] if 0 <= base_font_id < len(font_list) else (font_list[0] if font_list else None)
        hyperlink_font_id = _ooxml_ensure_hyperlink_font(fonts, base_font)
        font_list = _ooxml_child_elements(fonts, "font")
        style_map[base_style_id] = _ooxml_ensure_hyperlink_xf(cell_xfs, base_xf, hyperlink_font_id)
        xf_list = _ooxml_child_elements(cell_xfs, "xf")

    xml_text = _ooxml_xml_bytes(root).decode("utf-8")
    xml_text = _ooxml_repair_ignorable_namespaces(xml_text, "styleSheet")
    return xml_text.encode("utf-8"), style_map


def _ooxml_patch_sheet_hyperlinks(
    sheet_xml: bytes,
    rels_xml: bytes | None,
    links_by_address: dict[str, str],
    style_by_address: dict[str, int] | None = None,
) -> tuple[bytes, bytes]:
    sheet_text = sheet_xml.decode("utf-8-sig")
    sheet_text = _ooxml_repair_ignorable_namespace_declarations(sheet_text)
    rel_prefix, sheet_text = _ooxml_relationship_prefix(sheet_text)
    if rels_xml:
        rels_root = ET.fromstring(rels_xml)
    else:
        rels_root = ET.Element(_ooxml_qname(OOXML_PACKAGE_REL_NS, "Relationships"))

    sheet_text, removed_relationship_ids = _ooxml_remove_sheet_hyperlinks(
        sheet_text,
        set(links_by_address),
    )

    for rel in list(rels_root):
        if rel.attrib.get("Id") in removed_relationship_ids and rel.attrib.get("Type") == OOXML_HYPERLINK_REL_TYPE:
            rels_root.remove(rel)

    rel_ids_by_address: dict[str, str] = {}
    for address, target in links_by_address.items():
        rel_id = _ooxml_next_relationship_id(rels_root)
        rel_ids_by_address[address] = rel_id
        ET.SubElement(
            rels_root,
            _ooxml_qname(OOXML_PACKAGE_REL_NS, "Relationship"),
            {
                "Id": rel_id,
                "Type": OOXML_HYPERLINK_REL_TYPE,
                "Target": target,
                "TargetMode": "External",
            },
        )
    sheet_text = _ooxml_add_sheet_hyperlinks(sheet_text, rel_ids_by_address, rel_prefix)
    sheet_text = _ooxml_set_cell_style_ids(sheet_text, style_by_address or {})
    sheet_text = _ooxml_repair_ignorable_namespace_declarations(sheet_text)
    return sheet_text.encode("utf-8"), _ooxml_xml_bytes(rels_root)


def _replace_ooxml_zip_entries(workbook_path: Path, replacements: dict[str, bytes]) -> None:
    handle, temp_name = tempfile.mkstemp(
        prefix=f"{workbook_path.stem}.",
        suffix=".tmp",
        dir=str(workbook_path.parent),
    )
    os.close(handle)
    temp_path = Path(temp_name)
    try:
        with zipfile.ZipFile(workbook_path, "r") as source, zipfile.ZipFile(
            temp_path,
            "w",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=6,
        ) as target:
            existing = {info.filename for info in source.infolist()}
            for info in source.infolist():
                payload = replacements.get(info.filename)
                if payload is None:
                    payload = source.read(info.filename)
                target.writestr(info, payload)
            for filename, payload in replacements.items():
                if filename not in existing:
                    target.writestr(filename, payload)
        os.replace(temp_path, workbook_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def write_links_with_ooxml(
    workbook_path: Path,
    cells: list[WorkbookOrderCell],
    results: dict[str, PurchaseLinkResult],
) -> int:
    links_by_sheet: dict[str, dict[str, str]] = {}
    for cell_info in cells:
        result = results.get(cell_info.order_name)
        if result is None or result.status != "linked" or not result.url:
            continue
        links_by_sheet.setdefault(cell_info.sheet, {})[cell_info.address] = result.url
    if not links_by_sheet:
        return 0

    count = 0
    with zipfile.ZipFile(workbook_path, "r") as workbook_zip:
        sheet_map = _ooxml_workbook_sheet_map(workbook_zip)
        workbook_names = set(workbook_zip.namelist())
        sheet_payloads: list[tuple[str, bytes, bytes | None, dict[str, str], dict[str, int]]] = []
        base_style_ids: set[int] = set()
        replacements: dict[str, bytes] = {}
        for sheet_name, links_by_address in links_by_sheet.items():
            sheet_xml_path = sheet_map.get(sheet_name)
            if not sheet_xml_path:
                continue
            rels_path = _ooxml_sheet_rels_path(sheet_xml_path)
            rels_xml = workbook_zip.read(rels_path) if rels_path in workbook_names else None
            sheet_xml = workbook_zip.read(sheet_xml_path)
            style_ids = _ooxml_collect_cell_style_ids(
                sheet_xml.decode("utf-8-sig"),
                set(links_by_address),
            )
            base_style_ids.update(style_ids.values())
            sheet_payloads.append((sheet_xml_path, sheet_xml, rels_xml, links_by_address, style_ids))

        styles_by_base_id: dict[int, int] = {}
        if base_style_ids and "xl/styles.xml" in workbook_names:
            patched_styles, styles_by_base_id = _ooxml_patch_styles_for_hyperlinks(
                workbook_zip.read("xl/styles.xml"),
                base_style_ids,
            )
            replacements["xl/styles.xml"] = patched_styles

        for sheet_xml_path, sheet_xml, rels_xml, links_by_address, style_ids in sheet_payloads:
            sheet_xml, new_rels_xml = _ooxml_patch_sheet_hyperlinks(
                sheet_xml,
                rels_xml,
                links_by_address,
                {
                    address: styles_by_base_id.get(base_style_id, base_style_id)
                    for address, base_style_id in style_ids.items()
                },
            )
            rels_path = _ooxml_sheet_rels_path(sheet_xml_path)
            replacements[sheet_xml_path] = sheet_xml
            replacements[rels_path] = new_rels_xml
            count += len(links_by_address)

    if not replacements:
        return 0
    _replace_ooxml_zip_entries(workbook_path, replacements)
    return count


def write_links_with_live_workbook(
    access: WorkbookAccessContext, cells: list[WorkbookOrderCell],
    results: dict[str, PurchaseLinkResult],
    save_debounce_seconds: float,
) -> int:
    workbook = access.workbook
    if workbook is None:
        raise WorkbookAccessError(access, "No live Excel workbook is attached.")
    if safe_excel_bool_property(workbook, "ReadOnly"):
        raise WorkbookAccessError(access, "The workbook is open read-only in Excel.")
    previous_alerts = None
    application = access.application or workbook.Application
    try:
        try:
            previous_alerts = application.DisplayAlerts
            application.DisplayAlerts = False
        except Exception:
            previous_alerts = None
        count = apply_links_to_workbook(workbook, cells, results)
        if count > 0:
            save_workbook_with_suppressed_events(workbook, save_debounce_seconds)
        return count
    finally:
        if previous_alerts is not None:
            try:
                application.DisplayAlerts = previous_alerts
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Batch resolution
# ---------------------------------------------------------------------------

def resolve_orders(
    ref_values: list[str],
    client: OdooClient,
    lookup_mode: str = LOOKUP_MODE_PARTNER_REF,
    url_builder: Any = None,
    global_search_on_not_found: bool = False,
) -> dict[str, PurchaseLinkResult]:
    unique = list(dict.fromkeys(ref_values))
    results = resolve_orders_exact_batch(unique, client, lookup_mode)

    if global_search_on_not_found:
        missing_after_exact = [ref for ref in unique if ref not in results]
        if missing_after_exact:
            results.update(resolve_global_fast_exact_refs(missing_after_exact, client))

    remaining = [ref for ref in unique if ref not in results]
    thread_local = threading.local()

    def _thread_client() -> OdooClient:
        cached = getattr(thread_local, "client", None)
        if cached is None:
            cached = OdooClient(client.url, client.db, client.login, client.api_key)
            cached.authenticate()
            thread_local.client = cached
        return cached

    def _resolve_one(ref: str) -> tuple[str, PurchaseLinkResult]:
        thread_client = _thread_client()
        return ref, thread_client.resolve_purchase_ref(ref, lookup_mode=lookup_mode)

    max_workers = min(DEFAULT_ODOO_MAX_WORKERS, max(1, len(remaining)))
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(_resolve_one, ref): ref for ref in remaining}
        for future in as_completed(futures):
            ref, result = future.result()
            results[ref] = result

    if global_search_on_not_found:
        unresolved = [
            ref for ref in unique
            if results.get(ref, PurchaseLinkResult(status="not_found")).status == "not_found"
        ]
        if unresolved:
            global_contains = resolve_global_fast_contains_refs(unresolved, client)
            for ref, result in global_contains.items():
                if result.status == "linked":
                    results[ref] = result
    return results


def resolve_orders_exact_batch(
    ref_values: list[str],
    client: OdooClient,
    lookup_mode: str,
) -> dict[str, PurchaseLinkResult]:
    if client.__class__ is not OdooClient:
        return {}
    if not ref_values:
        return {}
    if lookup_mode == LOOKUP_MODE_COMMAND_REF:
        return _resolve_command_refs_exact_batch(ref_values, client)
    if lookup_mode == LOOKUP_MODE_TOTAL_AMOUNT:
        return _resolve_amounts_exact_batch(ref_values, client)
    if lookup_mode == LOOKUP_MODE_PARTNER_REF:
        return _resolve_text_refs_exact_batch(ref_values, client, "partner_ref")
    if lookup_mode == LOOKUP_MODE_GLOBAL_TEXT:
        return resolve_global_exact_refs(ref_values, client)
    return {}


def _resolve_text_refs_exact_batch(
    ref_values: list[str],
    client: OdooClient,
    field_name: str,
) -> dict[str, PurchaseLinkResult]:
    refs = [normalize_order(ref) for ref in ref_values if normalize_order(ref)]
    orders = client.search_purchase_orders_by_exact_values(field_name, list(dict.fromkeys(refs)))
    by_value: dict[str, dict] = {}
    for order in orders:
        key = str(order.get(field_name) or "").strip().casefold()
        if key and key not in by_value:
            by_value[key] = order
    results: dict[str, PurchaseLinkResult] = {}
    for ref in ref_values:
        key = normalize_order(ref).casefold()
        order = by_value.get(key)
        if order is None:
            continue
        results[ref] = client._linked_order_result(
            order,
            matched_field=f"exact:{field_name}",
            ref_value=str(order.get(field_name) or ""),
        )
    return results


def _resolve_command_refs_exact_batch(ref_values: list[str], client: OdooClient) -> dict[str, PurchaseLinkResult]:
    results = _resolve_text_refs_exact_batch(ref_values, client, "name")
    remaining = [ref for ref in ref_values if ref not in results]
    if remaining:
        results.update(_resolve_text_refs_exact_batch(remaining, client, "partner_ref"))
    return results


def _resolve_amounts_exact_batch(ref_values: list[str], client: OdooClient) -> dict[str, PurchaseLinkResult]:
    amount_by_ref: dict[str, float | None] = {ref: parse_amount(ref) for ref in ref_values}
    valid_amounts = [amount for amount in amount_by_ref.values() if amount is not None]
    results: dict[str, PurchaseLinkResult] = {
        ref: PurchaseLinkResult(
            status="invalid_amount",
            note="Reference could not be parsed as amount for amount_total lookup.",
        )
        for ref, amount in amount_by_ref.items()
        if amount is None
    }
    if not valid_amounts:
        return results
    orders = client.search_purchase_orders_by_exact_values("amount_total", list(dict.fromkeys(valid_amounts)))
    orders_by_amount: dict[float, list[dict]] = {}
    for order in orders:
        try:
            amount = float(order.get("amount_total") or 0.0)
        except Exception:
            continue
        orders_by_amount.setdefault(round(amount, 2), []).append(order)
    for ref, amount in amount_by_ref.items():
        if amount is None:
            continue
        candidates = orders_by_amount.get(round(amount, 2), [])
        if not candidates:
            continue
        best = client._pick_best_order_by_amount(amount, candidates)
        results[ref] = client._linked_order_result(
            best,
            matched_field="exact:amount_total",
            ref_value=str(best.get("amount_total") or ""),
        )
    return results


def _global_search_cache_key(client: OdooClient) -> tuple[str, str, str]:
    return (
        normalize_odoo_url(getattr(client, "url", "")),
        str(getattr(client, "db", "") or ""),
        str(getattr(client, "login", "") or ""),
    )


def _is_skippable_global_search_error(exc: Exception) -> bool:
    message = str(exc).casefold()
    transient_markers = (
        "temporarily overloaded",
        "connection slots",
        "connection reset",
        "connection refused",
        "connection timed out",
        "network error",
        "bad gateway",
        "gateway timeout",
        "service unavailable",
    )
    return not any(marker in message for marker in transient_markers)


def _global_model_candidates(client: OdooClient) -> list[str]:
    try:
        models = client._call(
            "ir.model",
            "search_read",
            [[]],
            {"fields": ["model", "transient"], "limit": 10000, "order": "model asc"},
        )
    except Exception:
        return list(GLOBAL_SEARCH_MODEL_ORDER)
    candidates: list[str] = []
    for item in models:
        if item.get("transient"):
            continue
        model = str(item.get("model") or "").strip()
        if model:
            candidates.append(model)
    return list(dict.fromkeys(list(GLOBAL_SEARCH_MODEL_ORDER) + candidates))


def _order_global_fields(model: str, fields: list[str]) -> list[str]:
    known = set(fields)
    priority = [field for field in GLOBAL_SEARCH_PRIORITY_FIELDS.get(model, ()) if field in known]
    rest = [field for field in fields if field not in priority]
    rest.sort(
        key=lambda field: (
            0 if any(token in field.casefold() for token in ("ref", "name", "origin", "number", "invoice", "facture", "x_studio")) else 1,
            field,
        )
    )
    return priority + rest


def _global_search_model_fields(client: OdooClient) -> dict[str, list[str]]:
    cache_key = _global_search_cache_key(client)
    cached = _GLOBAL_SEARCH_FIELDS_CACHE.get(cache_key)
    if cached is not None:
        return cached

    model_fields: dict[str, list[str]] = {}
    for model in _global_model_candidates(client):
        try:
            has_access = client._call(model, "check_access_rights", ["read"], {"raise_exception": False})
        except Exception as exc:
            if not _is_skippable_global_search_error(exc):
                raise
            continue
        if not has_access:
            continue
        try:
            fields = client._call(
                model,
                "fields_get",
                [],
                {"attributes": ["string", "type", "searchable", "store"]},
            )
        except Exception as exc:
            if not _is_skippable_global_search_error(exc):
                raise
            continue
        selected = [
            field_name
            for field_name, metadata in fields.items()
            if metadata.get("searchable", True)
            and metadata.get("type") in GLOBAL_SEARCH_TEXT_TYPES
            and field_name not in GLOBAL_SEARCH_EXCLUDED_FIELDS
        ]
        if selected:
            model_fields[model] = _order_global_fields(model, selected)
    _GLOBAL_SEARCH_FIELDS_CACHE[cache_key] = model_fields
    return model_fields


def _global_fast_model_fields(
    client: OdooClient,
    configured_fields: dict[str, tuple[str, ...]],
    cache_scope: str,
) -> dict[str, list[str]]:
    base_key = _global_search_cache_key(client)
    cache_key = (base_key[0], base_key[1], base_key[2], cache_scope)
    cached = _GLOBAL_FAST_FIELDS_CACHE.get(cache_key)
    if cached is not None:
        return cached

    model_fields: dict[str, list[str]] = {}
    for model, desired_fields in configured_fields.items():
        try:
            has_access = client._call(model, "check_access_rights", ["read"], {"raise_exception": False})
        except Exception as exc:
            if not _is_skippable_global_search_error(exc):
                raise
            continue
        if not has_access:
            continue
        try:
            fields = client._call(
                model,
                "fields_get",
                [],
                {"attributes": ["type", "searchable"]},
            )
        except Exception as exc:
            if not _is_skippable_global_search_error(exc):
                raise
            continue
        selected = [
            field_name
            for field_name in desired_fields
            if field_name in fields
            and fields[field_name].get("searchable", True)
            and fields[field_name].get("type") in GLOBAL_SEARCH_TEXT_TYPES
            and field_name not in GLOBAL_SEARCH_EXCLUDED_FIELDS
        ]
        if selected:
            model_fields[model] = selected
    _GLOBAL_FAST_FIELDS_CACHE[cache_key] = model_fields
    return model_fields


def _chunks(values: list[Any], size: int) -> list[list[Any]]:
    return [values[index:index + size] for index in range(0, len(values), size)]


def _or_domain(field_names: list[str], operator: str, value: Any) -> list[Any]:
    conditions = [(field_name, operator, value) for field_name in field_names]
    if len(conditions) <= 1:
        return conditions
    return ["|"] * (len(conditions) - 1) + conditions


def _record_text_value(value: Any) -> str:
    if value is False or value is None:
        return ""
    if isinstance(value, (list, tuple)):
        if len(value) >= 2:
            return str(value[1] or "")
        if value:
            return str(value[0] or "")
        return ""
    return str(value)


def _global_result(
    *,
    client: OdooClient,
    model: str,
    record: dict[str, Any],
    field_name: str,
    field_value: str,
    match_type: str,
    query: str,
) -> PurchaseLinkResult:
    record_id = int(record.get("id") or 0)
    display_name = str(record.get("display_name") or record.get("name") or "")
    matched_field = f"global_{match_type}:{model}.{field_name}"
    return PurchaseLinkResult(
        status="linked",
        source_model=model,
        matched_field=matched_field,
        record_id=record_id,
        record_name=display_name,
        ref_value=field_value,
        url=build_odoo_record_url(client.url, model, record_id),
        note=f"Matched by global Odoo search via {matched_field} using '{query}'.",
    )


def _find_exact_global_matches(
    refs_by_key: dict[str, str],
    record: dict[str, Any],
    fields: list[str],
) -> list[tuple[str, str, str]]:
    matches: list[tuple[str, str, str]] = []
    for field_name in fields:
        value = _record_text_value(record.get(field_name)).strip()
        if not value:
            continue
        ref = refs_by_key.get(value.casefold())
        if ref:
            matches.append((ref, field_name, value))
    return matches


def _find_contains_global_match(
    query: str,
    record: dict[str, Any],
    fields: list[str],
) -> tuple[str, str] | None:
    folded = query.casefold()
    for field_name in fields:
        value = _record_text_value(record.get(field_name)).strip()
        if value and folded in value.casefold():
            return field_name, value
    return None


def _ref_variants(ref: str) -> list[str]:
    raw = normalize_order(ref)
    candidates = [
        raw,
        raw.replace(" ", ""),
        re.sub(r"[\s\-_/]+", "", raw),
        re.sub(r"\W+", "", raw),
    ]
    variants: list[str] = []
    for candidate in candidates:
        candidate = candidate.strip()
        if len(candidate) >= 2 and candidate not in variants:
            variants.append(candidate)
    return variants


def _search_global_exact_field_group(
    client: OdooClient,
    model: str,
    field_group: list[str],
    refs: list[str],
) -> list[dict[str, Any]]:
    read_fields = list(dict.fromkeys(GENERIC_RECORD_FIELDS + field_group))
    return client._call(
        model,
        "search_read",
        [_or_domain(field_group, "in", refs)],
        {
            "fields": read_fields,
            "limit": max(GLOBAL_SEARCH_RECORD_LIMIT, len(refs) * 3),
            "order": "id desc",
        },
    )


def _search_global_contains_field_group(
    client: OdooClient,
    model: str,
    field_group: list[str],
    query: str,
) -> list[dict[str, Any]]:
    read_fields = list(dict.fromkeys(GENERIC_RECORD_FIELDS + field_group))
    return client._call(
        model,
        "search_read",
        [_or_domain(field_group, "ilike", query)],
        {
            "fields": read_fields,
            "limit": GLOBAL_SEARCH_CONTAINS_LIMIT,
            "order": "id desc",
        },
    )


def resolve_global_exact_refs(
    ref_values: list[str],
    client: OdooClient,
) -> dict[str, PurchaseLinkResult]:
    refs = list(dict.fromkeys(normalize_order(ref) for ref in ref_values if normalize_order(ref)))
    results: dict[str, PurchaseLinkResult] = {}
    if not refs:
        return results
    refs_by_key = {ref.casefold(): ref for ref in refs}
    model_fields = _global_search_model_fields(client)
    for model, fields in model_fields.items():
        unresolved = [ref for ref in refs if ref not in results]
        if not unresolved:
            break
        for field_group in _chunks(fields, GLOBAL_SEARCH_FIELD_GROUP_SIZE):
            for ref_chunk in _chunks(unresolved, GLOBAL_SEARCH_REF_CHUNK_SIZE):
                try:
                    records = _search_global_exact_field_group(client, model, field_group, ref_chunk)
                except Exception as exc:
                    if not _is_skippable_global_search_error(exc):
                        raise
                    for field_name in field_group:
                        try:
                            records = _search_global_exact_field_group(client, model, [field_name], ref_chunk)
                        except Exception as field_exc:
                            if not _is_skippable_global_search_error(field_exc):
                                raise
                            continue
                        for record in records:
                            for ref, matched_field, value in _find_exact_global_matches(refs_by_key, record, [field_name]):
                                if ref not in results:
                                    results[ref] = _global_result(
                                        client=client,
                                        model=model,
                                        record=record,
                                        field_name=matched_field,
                                        field_value=value,
                                        match_type="exact",
                                        query=ref,
                                    )
                    continue
                for record in records:
                    for ref, matched_field, value in _find_exact_global_matches(refs_by_key, record, field_group):
                        if ref not in results:
                            results[ref] = _global_result(
                                client=client,
                                model=model,
                                record=record,
                                field_name=matched_field,
                                field_value=value,
                                match_type="exact",
                                query=ref,
                            )
    return results


def resolve_global_fast_exact_refs(
    ref_values: list[str],
    client: OdooClient,
) -> dict[str, PurchaseLinkResult]:
    refs = list(dict.fromkeys(normalize_order(ref) for ref in ref_values if normalize_order(ref)))
    results: dict[str, PurchaseLinkResult] = {}
    if not refs:
        return results
    refs_by_key = {ref.casefold(): ref for ref in refs}
    model_fields = _global_fast_model_fields(client, GLOBAL_FAST_EXACT_FIELDS, "exact")
    for model, fields in model_fields.items():
        unresolved = [ref for ref in refs if ref not in results]
        if not unresolved:
            break
        for ref_chunk in _chunks(unresolved, GLOBAL_SEARCH_REF_CHUNK_SIZE):
            try:
                records = _search_global_exact_field_group(client, model, fields, ref_chunk)
            except Exception as exc:
                if not _is_skippable_global_search_error(exc):
                    raise
                for field_name in fields:
                    try:
                        records = _search_global_exact_field_group(client, model, [field_name], ref_chunk)
                    except Exception as field_exc:
                        if not _is_skippable_global_search_error(field_exc):
                            raise
                        continue
                    for record in records:
                        for ref, matched_field, value in _find_exact_global_matches(refs_by_key, record, [field_name]):
                            if ref not in results:
                                results[ref] = _global_result(
                                    client=client,
                                    model=model,
                                    record=record,
                                    field_name=matched_field,
                                    field_value=value,
                                    match_type="exact",
                                    query=ref,
                                )
                continue
            for record in records:
                for ref, matched_field, value in _find_exact_global_matches(refs_by_key, record, fields):
                    if ref not in results:
                        results[ref] = _global_result(
                            client=client,
                            model=model,
                            record=record,
                            field_name=matched_field,
                            field_value=value,
                            match_type="exact",
                            query=ref,
                        )
    return results


def resolve_global_contains_refs(
    ref_values: list[str],
    client: OdooClient,
) -> dict[str, PurchaseLinkResult]:
    refs = list(dict.fromkeys(normalize_order(ref) for ref in ref_values if normalize_order(ref)))
    results: dict[str, PurchaseLinkResult] = {}
    if not refs:
        return results
    model_fields = _global_search_model_fields(client)
    for ref in refs:
        for query in _ref_variants(ref):
            found = False
            for model, fields in model_fields.items():
                for field_group in _chunks(fields, GLOBAL_SEARCH_FIELD_GROUP_SIZE):
                    try:
                        records = _search_global_contains_field_group(client, model, field_group, query)
                    except Exception as exc:
                        if not _is_skippable_global_search_error(exc):
                            raise
                        continue
                    for record in records:
                        match = _find_contains_global_match(query, record, field_group)
                        if match is None:
                            continue
                        field_name, value = match
                        results[ref] = _global_result(
                            client=client,
                            model=model,
                            record=record,
                            field_name=field_name,
                            field_value=value,
                            match_type="contains",
                            query=query,
                        )
                        found = True
                        break
                    if found:
                        break
                if found:
                    break
            if found:
                break
    return results


def resolve_global_fast_contains_refs(
    ref_values: list[str],
    client: OdooClient,
) -> dict[str, PurchaseLinkResult]:
    refs = list(dict.fromkeys(normalize_order(ref) for ref in ref_values if normalize_order(ref)))
    results: dict[str, PurchaseLinkResult] = {}
    if not refs:
        return results
    model_fields = _global_fast_model_fields(client, GLOBAL_FAST_CONTAINS_FIELDS, "contains")
    for ref in refs:
        for query in _ref_variants(ref):
            found = False
            for model, fields in model_fields.items():
                try:
                    records = _search_global_contains_field_group(client, model, fields, query)
                except Exception as exc:
                    if not _is_skippable_global_search_error(exc):
                        raise
                    continue
                for record in records:
                    match = _find_contains_global_match(query, record, fields)
                    if match is None:
                        continue
                    field_name, value = match
                    results[ref] = _global_result(
                        client=client,
                        model=model,
                        record=record,
                        field_name=field_name,
                        field_value=value,
                        match_type="contains",
                        query=query,
                    )
                    found = True
                    break
                if found:
                    break
            if found:
                break
    return results


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def write_report(
    report_path: Path, cells: list[WorkbookOrderCell],
    results: dict[str, PurchaseLinkResult],
    test_orders: list[str] | None = None,
) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=REPORT_COLUMNS)
        writer.writeheader()
        if test_orders is not None:
            for ref in test_orders:
                r = results.get(ref, PurchaseLinkResult(status="unknown"))
                writer.writerow(_report_row("", "", ref, r))
            return
        for cell in cells:
            r = results.get(cell.order_name, PurchaseLinkResult(status="unknown"))
            writer.writerow(_report_row(cell.sheet, cell.address, cell.order_name, _decorate_result_for_cell(cell, r)))


def _report_row(sheet: str, cell: str, ref: str, r: PurchaseLinkResult) -> dict[str, Any]:
    return {
        "sheet": sheet, "cell": cell, "reference": ref,
        "status": r.status, "source_model": r.source_model,
        "matched_field": r.matched_field,
        "record_id": r.record_id or "", "record_name": r.record_name,
        "ref_value": r.ref_value, "state": r.state,
        "vendor": r.vendor, "amount": r.amount,
        "url": r.url, "note": r.note,
    }


def _decorate_result_for_cell(cell: WorkbookOrderCell, result: PurchaseLinkResult) -> PurchaseLinkResult:
    if not cell.fallback_used:
        return result
    note_suffix = f"Row fallback used after no match on {cell.fallback_from}."
    note = f"{result.note} {note_suffix}".strip()
    matched_field = result.matched_field
    if matched_field:
        matched_field = f"{matched_field};fallback"
    else:
        matched_field = "fallback"
    return replace(result, matched_field=matched_field, note=note)


def status_counts(results: dict[str, PurchaseLinkResult]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for r in results.values():
        counts[r.status] = counts.get(r.status, 0) + 1
    return counts


def workbook_access_message(access: WorkbookAccessContext, allow_open_workbook_update: bool) -> str:
    if access.status == "open_writable" and not allow_open_workbook_update:
        return "Workbook is open in Excel. Live update is disabled."
    if access.status == "open_read_only":
        return "Workbook is open in Excel as read-only."
    if access.status == "open_autosave":
        return "Workbook is open with AutoSave enabled."
    if access.status == "open_ambiguous_instance":
        return "Workbook is open in more than one Excel instance."
    if access.status == "unsupported_live_update":
        return access.details.strip() or "Live Excel monitoring is unavailable."
    return access.details or "Workbook is not ready for live update."


# ---------------------------------------------------------------------------
# process_workbook — main entry point used by the UI
# ---------------------------------------------------------------------------

def process_workbook(
    workbook_path: Path,
    odoo_url: str,
    odoo_db: str,
    odoo_login: str,
    odoo_api_key: str,
    record_url_example: str = "",
    report_path: Path | None = None,
    report_dir: Path | None = None,
    backup_dir: Path | None = None,
    write_report_file: bool = True,
    stable_backup_name: bool = False,
    apply: bool = False,
    visible_excel: bool = False,
    allow_open_workbook_update: bool = False,
    excel_session_backend: str = DEFAULT_EXCEL_SESSION_BACKEND,
    excel_save_debounce_seconds: int = DEFAULT_EXCEL_SAVE_DEBOUNCE_SECONDS,
    allow_live_update_with_autosave: bool = False,
    workbook_slot: str = "",
    performance_mode: str = PERFORMANCE_MODE_SILENT,
) -> WorkbookProcessSummary:
    workbook_path = workbook_path.expanduser().resolve()
    workbook_rule = workbook_rule_for_slot(workbook_slot, workbook_path)
    normalized_performance_mode = str(performance_mode or PERFORMANCE_MODE_SILENT).strip().casefold()
    live_mode = normalized_performance_mode == PERFORMANCE_MODE_LIVE
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    if not is_supported_workbook(workbook_path):
        raise ValueError(f"Unsupported workbook extension: {workbook_path.suffix}")

    with com_scope():
        if live_mode:
            access = attach_workbook_session(
                workbook_path,
                excel_session_backend=excel_session_backend,
                allow_live_update_with_autosave=allow_live_update_with_autosave,
                attach_objects=True,
            )
        else:
            access = WorkbookAccessContext(status="closed", workbook_path=workbook_path, backend="openpyxl")
        if access.workbook is not None and access.is_open:
            scan_result = collect_workbook_orders(
                access.workbook,
                workbook_rule,
            )
        else:
            scan_result = scan_workbook_orders_from_file(
                workbook_path,
                visible_excel,
                workbook_rule,
                allow_com_fallback=live_mode,
            )
        cells = scan_result.cells

        unique_refs = sorted({c.order_name for c in cells})
        resolved_report_path: Path | None = None
        if not unique_refs:
            if write_report_file:
                resolved_report_path = (report_path.resolve() if report_path
                                        else default_report_path(workbook_path, report_dir))
                write_report(resolved_report_path, cells, {})
            return WorkbookProcessSummary(
                workbook_path=workbook_path, report_path=resolved_report_path,
                backup_path=None, total_cells=len(cells), unique_orders=0,
                linked_count=0,
                status_counts=selected_status_counts([], {}, scan_result),
                workbook_state=access.status,
            )

        client = OdooClient(odoo_url, odoo_db, odoo_login, odoo_api_key)
        client.authenticate()
        results = resolve_orders(
            unique_refs,
            client,
            lookup_mode=workbook_rule.lookup_mode,
            global_search_on_not_found=workbook_rule.global_search_on_not_found,
        )
        selected_cells = select_cells_for_results(cells, results, workbook_rule)
        selected_counts = selected_status_counts(selected_cells, results, scan_result)

        if write_report_file:
            resolved_report_path = (report_path.resolve() if report_path
                                    else default_report_path(workbook_path, report_dir))
            write_report(resolved_report_path, selected_cells, results)

        backup_path_val: Path | None = None
        linked_count = 0
        live_update_used = False
        if apply and selected_cells:
            if access.status == "open_writable":
                if not allow_open_workbook_update:
                    raise WorkbookAccessError(access, workbook_access_message(access, allow_open_workbook_update))
                backup_path_val = backup_workbook(
                    workbook_path, backup_dir,
                    stable_backup_name=stable_backup_name,
                    workbook=access.workbook)
                linked_count = write_links_with_live_workbook(
                    access, selected_cells, results, excel_save_debounce_seconds)
                live_update_used = True
            elif access.status in LIVE_UPDATE_WAIT_STATES:
                raise WorkbookAccessError(access, workbook_access_message(access, allow_open_workbook_update))
            else:
                if linked_cell_count(selected_cells, results) > 0:
                    backup_path_val = backup_workbook(
                        workbook_path, backup_dir, stable_backup_name=stable_backup_name)
                    if workbook_path.suffix.lower() in {".xlsx", ".xlsm"}:
                        try:
                            linked_count = write_links_with_ooxml(
                                workbook_path,
                                selected_cells,
                                results,
                            )
                        except Exception:
                            if not live_mode:
                                raise
                            linked_count = write_links_with_excel(
                                workbook_path, selected_cells, results,
                                visible_excel, excel_save_debounce_seconds)
                    else:
                        linked_count = write_links_with_excel(
                            workbook_path, selected_cells, results,
                            visible_excel, excel_save_debounce_seconds)

        return WorkbookProcessSummary(
            workbook_path=workbook_path,
            report_path=resolved_report_path,
            backup_path=backup_path_val,
            total_cells=len(selected_cells),
            unique_orders=len(unique_refs),
            linked_count=linked_count,
            status_counts=selected_counts,
            workbook_state=access.status,
            live_update_used=live_update_used,
        )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--workbook", help="Path to an Excel workbook to process.")
    p.add_argument("--odoo-url", default=os.getenv("ODOO_URL", DEFAULT_ODOO_URL))
    p.add_argument("--odoo-db", default=os.getenv("ODOO_DB"))
    p.add_argument("--odoo-login", default=os.getenv("ODOO_LOGIN"))
    p.add_argument("--odoo-api-key", default=os.getenv("ODOO_API_KEY"))
    p.add_argument("--record-url-example", default="")
    p.add_argument("--report", default=None, help="CSV report path.")
    p.add_argument("--report-dir", default=None)
    p.add_argument("--backup-dir", default=None)
    p.add_argument("--test-orders", nargs="*", default=None,
                   help="Only resolve these references (no Excel scan).")
    p.add_argument("--apply", action="store_true",
                   help="Create backups and update the workbooks with hyperlinks.")
    p.add_argument("--visible-excel", action="store_true")
    p.add_argument("--excel-session-backend", default=DEFAULT_EXCEL_SESSION_BACKEND,
                   choices=("xlwings", "pywin32"))
    p.add_argument("--excel-save-debounce-seconds", type=int,
                   default=DEFAULT_EXCEL_SAVE_DEBOUNCE_SECONDS)
    p.add_argument("--allow-live-update-with-autosave", action="store_true")
    p.add_argument("--performance-mode", default=PERFORMANCE_MODE_SILENT,
                   choices=(PERFORMANCE_MODE_SILENT, PERFORMANCE_MODE_LIVE))
    p.add_argument("--prompt-secret", action="store_true")
    return p.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)

    missing = []
    if not args.odoo_db:
        missing.append("ODOO_DB / --odoo-db")
    if not args.odoo_login:
        missing.append("ODOO_LOGIN / --odoo-login")
    if not args.odoo_api_key:
        if args.prompt_secret:
            args.odoo_api_key = getpass.getpass("Odoo API key: ")
        if not args.odoo_api_key:
            missing.append("ODOO_API_KEY / --odoo-api-key")
    if missing:
        raise SystemExit("Missing: " + ", ".join(missing))

    client = OdooClient(args.odoo_url, args.odoo_db, args.odoo_login, args.odoo_api_key)
    client.authenticate()
    print(f"Authenticated to {args.odoo_url} as {args.odoo_login}")

    if args.test_orders is not None:
        test_refs = [r.strip() for r in args.test_orders if r.strip()]
        if not test_refs:
            raise SystemExit("--test-orders requires at least one reference.")
        results = resolve_orders(test_refs, client)
        for ref, result in results.items():
            print(f"\n  {ref}: {result.status}")
            if result.record_name:
                print(f"    Record: {result.record_name} ({result.source_model})")
            if result.url:
                print(f"    URL: {result.url}")
        return 0

    if not args.workbook:
        raise SystemExit("--workbook is required when not using --test-orders.")

    workbook_path = Path(args.workbook).expanduser().resolve()
    report_path = Path(args.report).resolve() if args.report else None
    report_dir = Path(args.report_dir).expanduser().resolve() if args.report_dir else None
    backup_dir = Path(args.backup_dir).expanduser().resolve() if args.backup_dir else None

    summary = process_workbook(
        workbook_path=workbook_path,
        odoo_url=args.odoo_url,
        odoo_db=args.odoo_db,
        odoo_login=args.odoo_login,
        odoo_api_key=args.odoo_api_key,
        record_url_example=args.record_url_example,
        report_path=report_path,
        report_dir=report_dir,
        backup_dir=backup_dir,
        apply=args.apply,
        visible_excel=args.visible_excel,
        excel_session_backend=args.excel_session_backend,
        excel_save_debounce_seconds=args.excel_save_debounce_seconds,
        allow_live_update_with_autosave=args.allow_live_update_with_autosave,
        performance_mode=args.performance_mode,
    )
    print(f"Found {summary.total_cells} cells, {summary.unique_orders} unique references.")
    print("Results: " + ", ".join(f"{k}={v}" for k, v in sorted(summary.status_counts.items())))
    if summary.report_path:
        print(f"Report: {summary.report_path}")
    if not args.apply:
        print("Dry-run. Re-run with --apply to update the workbook.")
        return 0
    if summary.backup_path:
        print(f"Backup: {summary.backup_path}")
    print(f"Hyperlinks added: {summary.linked_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
