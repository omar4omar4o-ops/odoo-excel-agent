import unittest
import xmlrpc.client
from pathlib import Path
from unittest.mock import patch
from typing import Any
from zipfile import ZipFile

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:  # pragma: no cover
    Workbook = None
    Table = None
    TableStyleInfo = None

from link_odoo_vendor_bills import (
    ALL_HEADER_VARIANTS,
    ETRANGER_TOTAL_HEADER_VARIANTS,
    LOCAL_COMMAND_HEADER_VARIANTS,
    LOCAL_HEADER_VARIANTS,
    LOOKUP_MODE_COMMAND_REF,
    LOOKUP_MODE_PARTNER_REF,
    LOOKUP_MODE_TOTAL_AMOUNT,
    OdooClient,
    PurchaseLinkResult,
    WORKBOOK_SLOT_ACHATS_ETRANGER,
    WORKBOOK_SLOT_SELLER_PREVIOUS,
    WorkbookOrderCell,
    _GLOBAL_SEARCH_FIELDS_CACHE,
    _ooxml_patch_sheet_hyperlinks,
    apply_links_to_workbook,
    build_backup_path,
    build_odoo_record_url,
    explain_odoo_exception,
    prepare_backup_dir_layout,
    resolve_global_exact_refs,
    resolve_orders,
    scan_workbook_orders_from_file,
    select_cells_for_results,
    validate_odoo_settings,
    write_links_with_ooxml,
    write_links_with_openpyxl,
    workbook_rule_for_path,
    workbook_rule_for_slot,
)


def make_order(
    order_id: int,
    *,
    name: str,
    partner_ref: str,
    amount_total: float,
) -> dict:
    return {
        "id": order_id,
        "name": name,
        "partner_ref": partner_ref,
        "partner_id": [11, "Vendor X"],
        "state": "purchase",
        "amount_total": amount_total,
    }


class RuleSelectionTests(unittest.TestCase):
    def test_local_workbook_uses_name_lookup_with_row_fallback_headers(self) -> None:
        rule = workbook_rule_for_path(Path(r"C:\tmp\EXCEL FACTURE ACHATS LOCAL.xlsx"))
        self.assertEqual(rule.lookup_mode, LOOKUP_MODE_COMMAND_REF)
        self.assertEqual(rule.headers, LOCAL_HEADER_VARIANTS | LOCAL_COMMAND_HEADER_VARIANTS)
        self.assertTrue(rule.row_fallback_on_not_found)
        self.assertTrue(rule.global_search_on_not_found)

    def test_etranger_workbook_uses_n_commande_lookup(self) -> None:
        rule = workbook_rule_for_path(Path(r"C:\tmp\TRACKING ACHATS ETRANGER (1).xlsx"))
        self.assertEqual(rule.lookup_mode, LOOKUP_MODE_COMMAND_REF)
        self.assertEqual(rule.headers, LOCAL_COMMAND_HEADER_VARIANTS)
        self.assertTrue(rule.global_search_on_not_found)

    def test_etranger_slot_forces_n_commande_lookup_for_renamed_workbook(self) -> None:
        rule = workbook_rule_for_slot(WORKBOOK_SLOT_ACHATS_ETRANGER, Path(r"C:\tmp\Renamed Copy.xlsx"))
        self.assertEqual(rule.lookup_mode, LOOKUP_MODE_COMMAND_REF)
        self.assertEqual(rule.headers, LOCAL_COMMAND_HEADER_VARIANTS)
        self.assertFalse(rule.row_fallback_on_not_found)
        self.assertTrue(rule.global_search_on_not_found)

    def test_seller_previous_slot_forces_n_commande_lookup(self) -> None:
        rule = workbook_rule_for_slot(WORKBOOK_SLOT_SELLER_PREVIOUS, Path(r"C:\tmp\L'ETAT DES COMMANDES.xlsx"))
        self.assertEqual(rule.lookup_mode, LOOKUP_MODE_COMMAND_REF)
        self.assertEqual(rule.headers, LOCAL_COMMAND_HEADER_VARIANTS)
        self.assertFalse(rule.row_fallback_on_not_found)
        self.assertTrue(rule.global_search_on_not_found)

    def test_other_workbooks_keep_legacy_lookup(self) -> None:
        rule = workbook_rule_for_path(Path(r"C:\tmp\Anything Else.xlsx"))
        self.assertEqual(rule.lookup_mode, LOOKUP_MODE_COMMAND_REF)
        self.assertEqual(rule.headers, ALL_HEADER_VARIANTS)


class CommandRefClient(OdooClient):
    def __init__(self) -> None:
        super().__init__("https://sphe.cloudoo.ma", "db", "login", "key")
        self.name_calls: list[tuple[str, str]] = []
        self.partner_ref_calls: list[tuple[str, str]] = []

    def search_purchase_orders_by_name(self, ref: str, operator: str) -> list[dict]:
        self.name_calls.append((ref, operator))
        if operator == "=":
            return []
        return [make_order(66, name="PO00066", partner_ref="FA356409", amount_total=120.0)]

    def search_purchase_orders_by_partner_ref(self, ref: str, operator: str) -> list[dict]:
        self.partner_ref_calls.append((ref, operator))
        return []


class CommandRefPartnerFallbackClient(OdooClient):
    def __init__(self) -> None:
        super().__init__("https://sphe.cloudoo.ma", "db", "login", "key")
        self.name_calls: list[tuple[str, str]] = []
        self.partner_ref_calls: list[tuple[str, str]] = []

    def search_purchase_orders_by_name(self, ref: str, operator: str) -> list[dict]:
        self.name_calls.append((ref, operator))
        return []

    def search_purchase_orders_by_partner_ref(self, ref: str, operator: str) -> list[dict]:
        self.partner_ref_calls.append((ref, operator))
        if operator == "=":
            return []
        return [make_order(77, name="PO00077", partner_ref="FA202603527", amount_total=460.01)]


class TotalClient(OdooClient):
    def __init__(self) -> None:
        super().__init__("https://sphe.cloudoo.ma", "db", "login", "key")
        self.exact_calls: list[float] = []
        self.range_calls: list[tuple[float, float]] = []

    def search_purchase_orders_by_amount_exact(self, amount: float) -> list[dict]:
        self.exact_calls.append(amount)
        return []

    def search_purchase_orders_by_amount_range(self, minimum: float, maximum: float) -> list[dict]:
        self.range_calls.append((minimum, maximum))
        return [
            make_order(310, name="PO00310", partner_ref="REF-A", amount_total=8753.76),
            make_order(309, name="PO00309", partner_ref="REF-B", amount_total=8753.76),
        ]


class TotalToleranceFallbackClient(OdooClient):
    def __init__(self) -> None:
        super().__init__("https://sphe.cloudoo.ma", "db", "login", "key")
        self.exact_calls: list[float] = []
        self.range_calls: list[tuple[float, float]] = []

    def search_purchase_orders_by_amount_exact(self, amount: float) -> list[dict]:
        self.exact_calls.append(amount)
        return []

    def search_purchase_orders_by_amount_range(self, minimum: float, maximum: float) -> list[dict]:
        self.range_calls.append((minimum, maximum))
        if round(maximum - minimum, 2) <= 0.02:
            return []
        return [make_order(401, name="PO00401", partner_ref="REF-401", amount_total=100.04)]


class FailingCommandRefClient(OdooClient):
    def __init__(self) -> None:
        super().__init__("https://sphe.cloudoo.ma", "db", "login", "key")

    def search_purchase_orders_by_name(self, ref: str, operator: str) -> list[dict]:
        raise RuntimeError("Odoo is temporarily overloaded")


class FailingTotalClient(OdooClient):
    def __init__(self) -> None:
        super().__init__("https://sphe.cloudoo.ma", "db", "login", "key")

    def search_purchase_orders_by_amount_exact(self, amount: float) -> list[dict]:
        raise RuntimeError("Odoo is temporarily overloaded")


class GlobalSearchFakeClient:
    def __init__(self) -> None:
        self.url = "https://sphe.cloudoo.ma"
        self.db = "db"
        self.login = "login"
        self.api_key = "key"
        self.calls: list[tuple[str, str]] = []

    def _call(self, model: str, method: str, args: list, kwargs: dict | None = None) -> object:
        self.calls.append((model, method))
        if model == "ir.model":
            raise RuntimeError("Access denied to ir.model")
        if method == "check_access_rights":
            return model == "account.move"
        if method == "fields_get":
            return {
                "ref": {"type": "char", "searchable": True},
                "payment_reference": {"type": "char", "searchable": True},
                "display_name": {"type": "char", "searchable": True},
                "amount_total": {"type": "float", "searchable": True},
            }
        if method == "search_read":
            domain_text = str(args[0])
            if "FA2026000045" in domain_text:
                return [
                    {
                        "id": 2310,
                        "display_name": "FACT/2026/0045",
                        "ref": "FA2026000045",
                        "payment_reference": False,
                    }
                ]
            return []
        raise AssertionError(f"Unexpected Odoo call: {model}.{method}")


class ResolveModeTests(unittest.TestCase):
    def setUp(self) -> None:
        _GLOBAL_SEARCH_FIELDS_CACHE.clear()

    def test_command_ref_lookup_uses_name(self) -> None:
        client = CommandRefClient()
        result = client.resolve_purchase_ref("PO000", lookup_mode=LOOKUP_MODE_COMMAND_REF)
        self.assertEqual(client.name_calls, [("PO000", "="), ("PO000", "ilike")])
        self.assertEqual(client.partner_ref_calls, [("PO000", "=")])
        self.assertEqual(result.status, "linked")
        self.assertEqual(result.matched_field, "contains:name")
        self.assertEqual(result.record_id, 66)
        self.assertEqual(result.url, "https://sphe.cloudoo.ma/odoo/purchase/66")
        self.assertEqual(result.ref_value, "PO00066")

    def test_command_ref_lookup_falls_back_to_partner_ref(self) -> None:
        client = CommandRefPartnerFallbackClient()
        result = client.resolve_purchase_ref("FA202603527", lookup_mode=LOOKUP_MODE_COMMAND_REF)
        self.assertEqual(result.status, "linked")
        self.assertEqual(result.matched_field, "exact:partner_ref")
        self.assertEqual(result.record_id, 77)
        self.assertEqual(result.url, "https://sphe.cloudoo.ma/odoo/purchase/77")
        self.assertEqual(client.name_calls, [("FA202603527", "="), ("FA202603527", "ilike")])
        self.assertEqual(
            client.partner_ref_calls,
            [("FA202603527", "="), ("FA202603527", "ilike")],
        )

    def test_total_lookup_uses_range_and_selects_latest(self) -> None:
        client = TotalClient()
        result = client.resolve_purchase_ref("8753.76", lookup_mode=LOOKUP_MODE_TOTAL_AMOUNT)
        self.assertEqual(result.status, "linked")
        self.assertEqual(result.matched_field, "range:amount_total+-0.01")
        self.assertEqual(result.record_id, 310)
        self.assertEqual(result.url, "https://sphe.cloudoo.ma/odoo/purchase/310")
        self.assertEqual(len(client.exact_calls), 1)
        self.assertEqual(len(client.range_calls), 1)
        self.assertAlmostEqual(client.exact_calls[0], 8753.76, places=2)
        self.assertAlmostEqual(client.range_calls[0][0], 8753.75, places=2)
        self.assertAlmostEqual(client.range_calls[0][1], 8753.77, places=2)

    def test_total_lookup_tries_wider_tolerance_when_needed(self) -> None:
        client = TotalToleranceFallbackClient()
        result = client.resolve_purchase_ref("100.00", lookup_mode=LOOKUP_MODE_TOTAL_AMOUNT)
        self.assertEqual(result.status, "linked")
        self.assertEqual(result.matched_field, "range:amount_total+-0.05")
        self.assertEqual(result.record_id, 401)
        self.assertEqual(len(client.exact_calls), 1)
        self.assertGreaterEqual(len(client.range_calls), 2)

    def test_total_lookup_invalid_amount(self) -> None:
        client = TotalClient()
        result = client.resolve_purchase_ref("NOT-A-NUMBER", lookup_mode=LOOKUP_MODE_TOTAL_AMOUNT)
        self.assertEqual(result.status, "invalid_amount")
        self.assertEqual(client.exact_calls, [])
        self.assertEqual(client.range_calls, [])

    def test_command_ref_lookup_propagates_odoo_errors(self) -> None:
        client = FailingCommandRefClient()
        with self.assertRaisesRegex(RuntimeError, "temporarily overloaded"):
            client.resolve_purchase_ref("FA202603527", lookup_mode=LOOKUP_MODE_COMMAND_REF)

    def test_total_lookup_propagates_odoo_errors(self) -> None:
        client = FailingTotalClient()
        with self.assertRaisesRegex(RuntimeError, "temporarily overloaded"):
            client.resolve_purchase_ref("8753.76", lookup_mode=LOOKUP_MODE_TOTAL_AMOUNT)

    def test_build_odoo_record_url_uses_purchase_shortcut_and_generic_form_url(self) -> None:
        self.assertEqual(
            build_odoo_record_url("https://sphe.cloudoo.ma", "purchase.order", 66),
            "https://sphe.cloudoo.ma/odoo/purchase/66",
        )
        self.assertEqual(
            build_odoo_record_url("https://sphe.cloudoo.ma", "account.move", 2310),
            "https://sphe.cloudoo.ma/web#id=2310&model=account.move&view_type=form",
        )

    def test_global_exact_search_finds_account_move_reference(self) -> None:
        client = GlobalSearchFakeClient()
        results = resolve_global_exact_refs(["FA2026000045"], client)  # type: ignore[arg-type]
        result = results["FA2026000045"]
        self.assertEqual(result.status, "linked")
        self.assertEqual(result.source_model, "account.move")
        self.assertEqual(result.matched_field, "global_exact:account.move.ref")
        self.assertEqual(result.record_id, 2310)
        self.assertEqual(result.url, "https://sphe.cloudoo.ma/web#id=2310&model=account.move&view_type=form")

    def test_resolve_orders_can_fallback_to_global_search(self) -> None:
        client = GlobalSearchFakeClient()
        results = resolve_orders(
            ["FA2026000045"],
            client,  # type: ignore[arg-type]
            lookup_mode=LOOKUP_MODE_COMMAND_REF,
            global_search_on_not_found=True,
        )
        self.assertEqual(results["FA2026000045"].status, "linked")
        self.assertEqual(results["FA2026000045"].source_model, "account.move")
        self.assertNotIn(("ir.model", "search_read"), client.calls)

    def test_etranger_n_commande_can_fallback_to_global_search(self) -> None:
        client = GlobalSearchFakeClient()
        results = resolve_orders(
            ["FA2026000045"],
            client,  # type: ignore[arg-type]
            lookup_mode=LOOKUP_MODE_COMMAND_REF,
            global_search_on_not_found=workbook_rule_for_slot(
                WORKBOOK_SLOT_ACHATS_ETRANGER,
                Path(r"C:\tmp\TRACKING ACHATS ETRANGER (1).xlsx"),
            ).global_search_on_not_found,
        )
        self.assertEqual(results["FA2026000045"].status, "linked")
        self.assertEqual(results["FA2026000045"].source_model, "account.move")

    def test_seller_previous_n_commande_can_fallback_to_global_search(self) -> None:
        client = GlobalSearchFakeClient()
        results = resolve_orders(
            ["FA2026000045"],
            client,  # type: ignore[arg-type]
            lookup_mode=LOOKUP_MODE_COMMAND_REF,
            global_search_on_not_found=workbook_rule_for_slot(
                WORKBOOK_SLOT_SELLER_PREVIOUS,
                Path(r"C:\tmp\L'ETAT DES COMMANDES.xlsx"),
            ).global_search_on_not_found,
        )
        self.assertEqual(results["FA2026000045"].status, "linked")
        self.assertEqual(results["FA2026000045"].source_model, "account.move")

    def test_global_fallback_updates_not_found_note_when_everything_fails(self) -> None:
        class AlwaysMissingClient(OdooClient):
            def __init__(self, url: str = "https://sphe.cloudoo.ma", db: str = "db", login: str = "login", api_key: str = "key") -> None:
                super().__init__(url, db, login, api_key)

            def authenticate(self) -> int:
                self.uid = 1
                return 1

            def search_purchase_orders_by_exact_values(self, field_name: str, values: list[str]) -> list[dict]:
                return []

            def search_purchase_orders_by_name(self, ref: str, operator: str) -> list[dict]:
                return []

            def search_purchase_orders_by_partner_ref(self, ref: str, operator: str) -> list[dict]:
                return []

            def _call(self, model: str, method: str, args: list, kwargs: dict | None = None) -> Any:
                if model == "ir.model":
                    return [{"model": "account.move", "name": "Journal Entry"}]
                if method == "fields_get":
                    return {"ref": {"type": "char", "searchable": True}}
                if method == "search_read":
                    return []
                if method == "check_access_rights":
                    return True
                raise AssertionError(f"Unexpected call: {model}.{method}")

        with patch("link_odoo_vendor_bills.OdooClient", AlwaysMissingClient):
            client = AlwaysMissingClient()
            results = resolve_orders(
                ["NO-SUCH-REF"],
                client,
                lookup_mode=LOOKUP_MODE_COMMAND_REF,
                global_search_on_not_found=True,
            )
        self.assertEqual(results["NO-SUCH-REF"].status, "not_found")
        self.assertIn("other accessible Odoo records", results["NO-SUCH-REF"].note)


class OdooSettingsValidationTests(unittest.TestCase):
    def test_validate_odoo_settings_rejects_url_in_db_field(self) -> None:
        with self.assertRaisesRegex(ValueError, "database must be the database name only"):
            validate_odoo_settings(
                "https://sphe.cloudoo.ma",
                "https://limewire.com/d/9nluz#2VJ4biiLRw",
                "user@example.com",
            )

    def test_validate_odoo_settings_rejects_path_in_odoo_url(self) -> None:
        with self.assertRaisesRegex(ValueError, "server root only"):
            validate_odoo_settings(
                "https://sphe.cloudoo.ma/odoo/purchase/1",
                "sphe.cloudoo.ma",
                "user@example.com",
            )

    def test_validate_odoo_settings_rejects_mismatched_record_url_host(self) -> None:
        with self.assertRaisesRegex(ValueError, "same host"):
            validate_odoo_settings(
                "https://sphe.cloudoo.ma",
                "sphe.cloudoo.ma",
                "user@example.com",
                "https://other.example.com/odoo/purchase/1",
            )

    def test_explain_odoo_exception_for_url_like_database(self) -> None:
        exc = xmlrpc.client.Fault(1, 'database "https://limewire.com/d/9nluz#2VJ4biiLRw" does not exist')
        message = str(
            explain_odoo_exception(
                exc,
                odoo_url="https://sphe.cloudoo.ma",
                odoo_db="https://limewire.com/d/9nluz#2VJ4biiLRw",
            )
        )
        self.assertIn("database field contains a URL", message)

    def test_explain_odoo_exception_for_connection_slots(self) -> None:
        exc = xmlrpc.client.Fault(
            1,
            "psycopg2.OperationalError: remaining connection slots are reserved for roles with the SUPERUSER attribute",
        )
        message = str(explain_odoo_exception(exc, odoo_url="https://sphe.cloudoo.ma", odoo_db="sphe.cloudoo.ma"))
        self.assertIn("temporarily overloaded", message)


class FakeApplication:
    def __init__(self) -> None:
        self.ScreenUpdating = True
        self.Calculation = -4105
        self.EnableEvents = True


class FakeCellHyperlinks:
    def __init__(self) -> None:
        self.deleted = False

    def Delete(self) -> None:
        self.deleted = True


class FakeCell:
    def __init__(self) -> None:
        self.Hyperlinks = FakeCellHyperlinks()


class FakeSheetHyperlinks:
    def __init__(self) -> None:
        self.added: list[dict] = []

    def Add(self, **kwargs) -> None:
        self.added.append(kwargs)


class FakeSheet:
    def __init__(self, name: str) -> None:
        self.Name = name
        self.Hyperlinks = FakeSheetHyperlinks()
        self._cells: dict[tuple[int, int], FakeCell] = {}

    def Cells(self, row: int, column: int) -> FakeCell:
        key = (row, column)
        if key not in self._cells:
            self._cells[key] = FakeCell()
        return self._cells[key]


class FakeWorkbook:
    def __init__(self, sheet: FakeSheet) -> None:
        self.Application = FakeApplication()
        self.Worksheets = [sheet]


class HyperlinkWriteTests(unittest.TestCase):
    def test_apply_links_keeps_original_cell_text(self) -> None:
        sheet = FakeSheet("Feuil1")
        workbook = FakeWorkbook(sheet)
        cells = [WorkbookOrderCell(sheet="Feuil1", row=9, column=3, address="C9", order_name="FA356409")]
        results = {
            "FA356409": PurchaseLinkResult(
                status="linked",
                source_model="purchase.order",
                record_id=66,
                url="https://sphe.cloudoo.ma/odoo/purchase/66",
            )
        }
        linked = apply_links_to_workbook(workbook, cells, results)
        self.assertEqual(linked, 1)
        self.assertEqual(len(sheet.Hyperlinks.added), 1)
        self.assertEqual(sheet.Hyperlinks.added[0]["TextToDisplay"], "FA356409")
        self.assertEqual(sheet.Hyperlinks.added[0]["Address"], "https://sphe.cloudoo.ma/odoo/purchase/66")

    @unittest.skipIf(Workbook is None, "openpyxl is not installed")
    def test_openpyxl_writer_keeps_original_cell_text(self) -> None:
        import tempfile
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "book.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Feuil1"
            sheet["C9"] = "FA356409"
            workbook.save(workbook_path)

            cells = [WorkbookOrderCell(sheet="Feuil1", row=9, column=3, address="C9", order_name="FA356409")]
            results = {
                "FA356409": PurchaseLinkResult(
                    status="linked",
                    source_model="purchase.order",
                    record_id=66,
                    url="https://sphe.cloudoo.ma/odoo/purchase/66",
                )
            }
            linked = write_links_with_openpyxl(workbook_path, cells, results)

            updated = load_workbook(workbook_path, read_only=False)
            try:
                updated_cell = updated["Feuil1"]["C9"]
                self.assertEqual(linked, 1)
                self.assertEqual(updated_cell.value, "FA356409")
                self.assertEqual(updated_cell.hyperlink.target, "https://sphe.cloudoo.ma/odoo/purchase/66")
            finally:
                updated.close()

    @unittest.skipIf(Workbook is None or Table is None, "openpyxl is not installed")
    def test_ooxml_writer_preserves_table_xml_and_writes_hyperlink(self) -> None:
        import tempfile
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "book.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Feuil1"
            sheet.append(["N°FACTURE", "Vendor", "Total"])
            sheet.append(["FA202603527", "TRUCK CENTER", 460.0])
            sheet.append(["FA202603528", "OTHER", 100.0])
            table = Table(displayName="Table1", ref="A1:C3")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
            workbook.save(workbook_path)
            workbook.close()

            with ZipFile(workbook_path, "r") as before_zip:
                table_xml_before = before_zip.read("xl/tables/table1.xml")

            cells = [WorkbookOrderCell(sheet="Feuil1", row=2, column=1, address="A2", order_name="FA202603527")]
            results = {
                "FA202603527": PurchaseLinkResult(
                    status="linked",
                    source_model="account.move",
                    record_id=2310,
                    url="https://sphe.cloudoo.ma/web#id=2310&model=account.move&view_type=form",
                )
            }
            linked = write_links_with_ooxml(workbook_path, cells, results)

            with ZipFile(workbook_path, "r") as after_zip:
                table_xml_after = after_zip.read("xl/tables/table1.xml")
                sheet_rels = after_zip.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf-8")

            updated = load_workbook(workbook_path, read_only=False)
            try:
                updated_cell = updated["Feuil1"]["A2"]
                self.assertEqual(linked, 1)
                self.assertEqual(table_xml_after, table_xml_before)
                self.assertIn("hyperlink", sheet_rels)
                self.assertEqual(updated_cell.value, "FA202603527")
                self.assertEqual(
                    updated_cell.hyperlink.target,
                    "https://sphe.cloudoo.ma/web#id=2310&model=account.move&view_type=form",
                )
                self.assertEqual(updated_cell.font.underline, "single")
                self.assertEqual(updated_cell.font.color.rgb, "FF0563C1")
            finally:
                updated.close()

    def test_ooxml_patch_repairs_ignorable_namespace_declarations(self) -> None:
        sheet_xml = (
            "<?xml version='1.0' encoding='utf-8'?>"
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:ns1="http://schemas.openxmlformats.org/markup-compatibility/2006" '
            'xmlns:ns2="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
            'ns1:Ignorable="x14ac xr xr2 xr3" ns2:uid="{00000000-0001-0000-0000-000000000000}">'
            '<sheetData><row r="1"><c r="A1" t="str"><v>FA202603527</v></c></row></sheetData>'
            '<tableParts count="1"><tablePart r:id="rId1"/></tableParts>'
            '</worksheet>'
        ).encode("utf-8")
        rels_xml = (
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/table" Target="../tables/table1.xml"/>'
            '</Relationships>'
        ).encode("utf-8")

        patched_sheet, patched_rels = _ooxml_patch_sheet_hyperlinks(
            sheet_xml,
            rels_xml,
            {"A1": "https://sphe.cloudoo.ma/web#id=1&model=account.move&view_type=form"},
        )
        patched_text = patched_sheet.decode("utf-8")
        self.assertIn('xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"', patched_text)
        self.assertIn('xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"', patched_text)
        self.assertIn('xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"', patched_text)
        self.assertIn('xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"', patched_text)
        self.assertIn("<hyperlinks>", patched_text)
        self.assertIn("TargetMode=\"External\"", patched_rels.decode("utf-8"))


@unittest.skipIf(Workbook is None, "openpyxl is not installed")
class LocalFallbackScanTests(unittest.TestCase):
    def test_etranger_slot_scan_reads_only_n_commande_and_ignores_mtt_de_facture(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "Renamed ACHATS ETRANGER.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Feuil1"
            sheet["C1"] = "N COMMANDE"
            sheet["H1"] = "MTT DE FACTURE"
            sheet["C2"] = "SA2026000002942"
            sheet["H2"] = 3542.87
            sheet["C3"] = "=SUBTOTAL(103,Table1[N COMMANDE])"
            workbook.save(workbook_path)

            rule = workbook_rule_for_slot(WORKBOOK_SLOT_ACHATS_ETRANGER, workbook_path)
            scan_result = scan_workbook_orders_from_file(workbook_path, visible_excel=False, workbook_rule=rule)

        self.assertEqual(scan_result.issue_code, "")
        self.assertEqual(len(scan_result.cells), 1)
        self.assertEqual(scan_result.cells[0].order_name, "SA2026000002942")
        self.assertEqual(scan_result.cells[0].address, "C2")
        self.assertEqual(scan_result.cells[0].header_name, "n commande")

    def test_seller_previous_slot_scan_reads_only_n_commande(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "L'ETAT DES COMMANDES.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Feuil1"
            sheet["B1"] = "N commandes"
            sheet["C1"] = "N°FACTURE"
            sheet["B2"] = "FA202603527"
            sheet["C2"] = "FA-IGNORED-BY-SELLER-SLOT"
            workbook.save(workbook_path)

            rule = workbook_rule_for_slot(WORKBOOK_SLOT_SELLER_PREVIOUS, workbook_path)
            scan_result = scan_workbook_orders_from_file(workbook_path, visible_excel=False, workbook_rule=rule)

        self.assertEqual(scan_result.issue_code, "")
        self.assertEqual(len(scan_result.cells), 1)
        self.assertEqual(scan_result.cells[0].order_name, "FA202603527")
        self.assertEqual(scan_result.cells[0].address, "B2")
        self.assertEqual(scan_result.cells[0].header_name, "n commandes")

    def test_local_scan_collects_both_headers_and_uses_secondary_when_primary_not_found(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "EXCEL FACTURE ACHATS LOCAL.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Feuil1"
            sheet["C1"] = "N°FACTURE"
            sheet["D1"] = "N commandes"
            sheet["C2"] = "FA-NOT-FOUND"
            sheet["D2"] = "FA202603527"
            workbook.save(workbook_path)

            rule = workbook_rule_for_path(workbook_path)
            scan_result = scan_workbook_orders_from_file(workbook_path, visible_excel=False, workbook_rule=rule)

        self.assertEqual(len(scan_result.cells), 2)
        self.assertEqual([cell.order_name for cell in scan_result.cells], ["FA-NOT-FOUND", "FA202603527"])
        results = {
            "FA-NOT-FOUND": PurchaseLinkResult(status="not_found"),
            "FA202603527": PurchaseLinkResult(status="linked", url="https://sphe.cloudoo.ma/odoo/purchase/77"),
        }
        selected = select_cells_for_results(scan_result.cells, results, rule)
        self.assertEqual(len(selected), 1)
        self.assertEqual(selected[0].order_name, "FA202603527")
        self.assertTrue(selected[0].fallback_used)
        self.assertEqual(selected[0].fallback_from, "FA-NOT-FOUND")

    def test_missing_required_header_is_reported(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "EXCEL FACTURE ACHATS LOCAL.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "Something else"
            workbook.save(workbook_path)

            rule = workbook_rule_for_path(workbook_path)
            scan_result = scan_workbook_orders_from_file(workbook_path, visible_excel=False, workbook_rule=rule)

        self.assertEqual(scan_result.cells, [])
        self.assertEqual(scan_result.issue_code, "missing_required_header")


class BackupLayoutTests(unittest.TestCase):
    def test_build_backup_path_uses_separate_subfolders(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook = root / "Book.xlsx"
            workbook.write_bytes(b"stub")

            stable_path = build_backup_path(workbook, root, stable_backup_name=True)
            run_path = build_backup_path(workbook, root, stable_backup_name=False)

        self.assertEqual(stable_path.parent.name, "original-snapshots")
        self.assertEqual(run_path.parent.name, "run-backups")

    def test_prepare_backup_dir_layout_moves_legacy_backup_files(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            legacy_original = root / "EXCEL FACTURE ACHATS LOCAL.1234567890.original.xlsx"
            legacy_backup = root / "EXCEL FACTURE ACHATS LOCAL.backup-20260509-155606.xlsx"
            report = root / "report.csv"
            legacy_original.write_bytes(b"orig")
            legacy_backup.write_bytes(b"run")
            report.write_text("ok", encoding="utf-8")

            layout = prepare_backup_dir_layout(root)

            self.assertFalse(legacy_original.exists())
            self.assertFalse(legacy_backup.exists())
            self.assertTrue((layout["originals_dir"] / legacy_original.name).exists())
            self.assertTrue((layout["run_dir"] / legacy_backup.name).exists())
            self.assertTrue(report.exists())


if __name__ == "__main__":
    unittest.main()
